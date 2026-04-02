"""
Backtest: align GRID round-by-round CS2 data with Polymarket historical odds.

For each matched GRID series + PM market pair:
  1. Walk through each round in the GRID data
  2. Find the PM price at that round's end timestamp
  3. Run the XGBoost model to get model's win probability
  4. Compute edge = model_prob - pm_price
  5. Track outcomes and simulate P&L

Usage: python backtest.py
"""

import json
import os
import re
import sys
from datetime import datetime, timedelta, timezone
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple

import config
from feature_extractor import extract_features_from_grid_state, features_to_dict
from model import MapWinModel

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_iso_duration(s: str) -> float:
    """Parse ISO 8601 duration like 'PT2M10.5S' into seconds."""
    if not s:
        return 0.0
    m = re.match(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:([\d.]+)S)?", s)
    if not m:
        return 0.0
    hours = int(m.group(1) or 0)
    mins = int(m.group(2) or 0)
    secs = float(m.group(3) or 0)
    return hours * 3600 + mins * 60 + secs


def parse_iso_ts(s: str) -> float:
    """Parse ISO 8601 timestamp to unix epoch seconds."""
    if not s:
        return 0.0
    # Handle Z and +00:00 variants
    s = s.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(s)
        return dt.timestamp()
    except Exception:
        return 0.0


def normalize_team(name: str) -> str:
    """Lowercase, strip common suffixes for fuzzy matching."""
    n = name.lower().strip()
    # Remove common suffixes/prefixes that differ between PM and GRID
    for suffix in [" esports", " esport", " gaming", " academy", " ac", " ks"]:
        if n.endswith(suffix):
            n = n[: -len(suffix)].strip()
    for prefix in ["team ", "ex-"]:
        if n.startswith(prefix):
            n = n[len(prefix):].strip()
    return n


def teams_match(pm_teams: List[str], grid_teams: List[str]) -> bool:
    """Check if PM and GRID team pairs refer to the same match.

    Uses progressively looser matching:
      1. Exact (case-insensitive)
      2. Normalized (strip suffixes)
      3. Substring containment
    """
    pm_lower = {t.lower() for t in pm_teams}
    grid_lower = {t.lower() for t in grid_teams}

    # Exact
    if pm_lower == grid_lower:
        return True

    # Normalized
    pm_norm = {normalize_team(t) for t in pm_teams}
    grid_norm = {normalize_team(t) for t in grid_teams}
    if pm_norm == grid_norm:
        return True

    # Substring: every PM team name is contained in some GRID name or vice versa
    def subset_match(set_a, set_b):
        for a in set_a:
            if not any(a in b or b in a for b in set_b):
                return False
        return True

    if subset_match(pm_norm, grid_norm) and subset_match(grid_norm, pm_norm):
        return True

    return False


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

@dataclass
class PMMarket:
    """One Polymarket Map Winner market with price history."""
    filename: str
    question: str
    outcomes: List[str]       # [team_a_name, team_b_name]
    date_str: str             # from filename, e.g. "2026-03-31"
    map_num: int              # 1, 2, 3
    volume: float             # market volume in USD
    # history for outcome_0 (first team listed)
    history_0: List[dict]     # [{t: epoch, p: prob}, ...]
    history_1: List[dict]


def load_pm_markets() -> List[PMMarket]:
    """Load all Map N Winner PM history files."""
    pm_dir = os.path.join(config.DATA_DIR, "pm_history")
    if not os.path.isdir(pm_dir):
        print(f"[WARN] PM history dir not found: {pm_dir}")
        return []

    markets = []
    for fname in os.listdir(pm_dir):
        if not fname.endswith(".json"):
            continue
        fpath = os.path.join(pm_dir, fname)
        try:
            with open(fpath) as f:
                d = json.load(f)
        except Exception:
            continue

        q = d.get("question", "")
        if "Map" not in q or "Winner" not in q:
            continue

        raw_o = d.get("outcomes", "[]")
        outcomes = json.loads(raw_o) if isinstance(raw_o, str) else raw_o
        if len(outcomes) < 2:
            continue

        date_m = re.search(r"(\d{4}-\d{2}-\d{2})", fname)
        date_str = date_m.group(1) if date_m else ""
        game_m = re.search(r"game(\d+)", fname)
        map_num = int(game_m.group(1)) if game_m else 0

        h = d.get("histories", {})
        h0 = h.get("outcome_0", {}).get("history", [])
        h1 = h.get("outcome_1", {}).get("history", [])

        if not h0 and not h1:
            continue

        volume = float(d.get("volume", 0) or 0)
        if volume < MIN_VOLUME:
            continue

        markets.append(PMMarket(
            filename=fname,
            question=q,
            outcomes=outcomes,
            date_str=date_str,
            map_num=map_num,
            volume=volume,
            history_0=h0,
            history_1=h1,
        ))

    return markets


@dataclass
class GRIDSeries:
    """One GRID series with metadata."""
    filename: str
    series_id: str
    teams: List[str]          # from _meta.teams_central
    scheduled: str            # ISO timestamp
    scheduled_date: str       # just the date part
    data: dict                # full series dict


def load_grid_series() -> List[GRIDSeries]:
    """Load all GRID series files."""
    series_dir = os.path.join(config.DATA_DIR, "series")
    if not os.path.isdir(series_dir):
        print(f"[WARN] Series dir not found: {series_dir}")
        return []

    series_list = []
    for fname in os.listdir(series_dir):
        if not fname.endswith(".json"):
            continue
        fpath = os.path.join(series_dir, fname)
        try:
            with open(fpath) as f:
                d = json.load(f)
        except Exception:
            continue

        meta = d.get("_meta", {})
        teams = meta.get("teams_central", [])
        sched = meta.get("scheduled", "")
        if not teams or not sched:
            continue

        series_list.append(GRIDSeries(
            filename=fname,
            series_id=str(d.get("id", fname.replace(".json", ""))),
            teams=teams,
            scheduled=sched,
            scheduled_date=sched[:10],
            data=d,
        ))

    return series_list


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

@dataclass
class MatchedPair:
    """A GRID game matched to a PM market."""
    grid_series: GRIDSeries
    game_idx: int             # 0-indexed game within series
    pm_market: PMMarket
    # Which PM outcome index corresponds to GRID team A (game_teams[0])
    pm_outcome_idx_for_team_a: int


def match_grid_to_pm(
    grid_series: List[GRIDSeries],
    pm_markets: List[PMMarket],
) -> List[MatchedPair]:
    """Match GRID series/games to PM Map Winner markets."""

    # Index GRID series by date for fast lookup
    grid_by_date: Dict[str, List[GRIDSeries]] = {}
    for gs in grid_series:
        grid_by_date.setdefault(gs.scheduled_date, []).append(gs)

    pairs = []
    for pm in pm_markets:
        if pm.map_num < 1:
            continue

        # Search GRID series on the PM date and +/- 1 day
        candidates = []
        for delta in [0, -1, 1]:
            try:
                d = datetime.strptime(pm.date_str, "%Y-%m-%d") + timedelta(days=delta)
                ds = d.strftime("%Y-%m-%d")
            except Exception:
                continue
            candidates.extend(grid_by_date.get(ds, []))

        for gs in candidates:
            if not teams_match(pm.outcomes, gs.teams):
                continue

            game_idx = pm.map_num - 1  # 0-indexed
            games = gs.data.get("games", [])
            if game_idx >= len(games):
                continue

            game = games[game_idx]
            if not game.get("finished"):
                continue
            segments = game.get("segments", [])
            if len(segments) < 3:
                continue  # Too few rounds to be interesting

            # Figure out which PM outcome corresponds to GRID team A
            game_teams = game.get("teams", [])
            if len(game_teams) < 2:
                continue
            grid_team_a = game_teams[0]["name"]

            # Match grid_team_a to PM outcomes
            pm_idx = -1
            for oi, oname in enumerate(pm.outcomes):
                if oname.lower() == grid_team_a.lower():
                    pm_idx = oi
                    break
                if normalize_team(oname) == normalize_team(grid_team_a):
                    pm_idx = oi
                    break
                if oname.lower() in grid_team_a.lower() or grid_team_a.lower() in oname.lower():
                    pm_idx = oi
                    break

            if pm_idx < 0:
                # Team A must be PM outcome 1's team (by elimination)
                # Actually we should check team B
                grid_team_b = game_teams[1]["name"]
                for oi, oname in enumerate(pm.outcomes):
                    if (oname.lower() == grid_team_b.lower()
                            or normalize_team(oname) == normalize_team(grid_team_b)
                            or oname.lower() in grid_team_b.lower()
                            or grid_team_b.lower() in oname.lower()):
                        pm_idx = 1 - oi  # team A is the other one
                        break

            if pm_idx < 0:
                continue  # Could not determine mapping

            pairs.append(MatchedPair(
                grid_series=gs,
                game_idx=game_idx,
                pm_market=pm,
                pm_outcome_idx_for_team_a=pm_idx,
            ))
            break  # Found a match for this PM market

    return pairs


# ---------------------------------------------------------------------------
# Timestamp alignment
# ---------------------------------------------------------------------------

def get_round_start_timestamps(game: dict) -> List[Tuple[int, float]]:
    """Return [(round_num, unix_timestamp_of_round_start), ...] for a game.

    Uses round START time so we compare model predictions against PM prices
    BEFORE the round resolves (no look-ahead bias from PM reacting to the
    round outcome).
    """
    segments = game.get("segments", [])
    results = []
    for seg in segments:
        if not seg.get("finished"):
            continue
        seq = seg.get("sequenceNumber", 0)
        started = parse_iso_ts(seg.get("startedAt", ""))
        if started > 0:
            results.append((seq, started))
    results.sort(key=lambda x: x[0])
    return results


def find_pm_price_at(history: List[dict], ts: float) -> Optional[float]:
    """Binary search for the PM price point with closest timestamp <= ts.

    Returns the price, or None if no data point exists before ts.
    """
    if not history:
        return None

    lo, hi = 0, len(history) - 1
    # If ts is before all history, return None
    if history[0]["t"] > ts:
        return None

    # Binary search for rightmost entry with t <= ts
    result_idx = 0
    while lo <= hi:
        mid = (lo + hi) // 2
        if history[mid]["t"] <= ts:
            result_idx = mid
            lo = mid + 1
        else:
            hi = mid - 1

    return history[result_idx]["p"]


# ---------------------------------------------------------------------------
# Backtest core
# ---------------------------------------------------------------------------

@dataclass
class RoundResult:
    """Result of one round in the backtest."""
    series_id: str
    game_idx: int
    round_num: int
    model_prob_a: float       # Model's P(team A wins map)
    pm_price_a: float         # PM price for team A winning
    edge: float               # model_prob_a - pm_price_a
    team_a: str
    team_b: str
    team_a_won_map: bool
    map_name: str


def run_backtest(pairs: List[MatchedPair], model: MapWinModel) -> List[RoundResult]:
    """Run the model on each round of each matched pair."""
    all_results = []

    for pair in pairs:
        gs = pair.grid_series
        game_idx = pair.game_idx
        pm = pair.pm_market

        game = gs.data["games"][game_idx]
        game_teams = game.get("teams", [])
        team_a = game_teams[0]["name"]
        team_b = game_teams[1]["name"]
        team_a_won = game_teams[0].get("won", False)
        map_name = (game.get("map") or {}).get("name", "unknown")

        # Get round-start timestamps (avoids look-ahead from PM reacting)
        round_starts = get_round_start_timestamps(game)
        if not round_starts:
            continue

        # Pick the right PM history based on outcome mapping
        if pair.pm_outcome_idx_for_team_a == 0:
            pm_hist = pm.history_0
        else:
            pm_hist = pm.history_1

        # Sort PM history by timestamp
        pm_hist_sorted = sorted(pm_hist, key=lambda x: x["t"])

        # Extract features for all rounds
        try:
            features_list = extract_features_from_grid_state(
                gs.data, game_idx=game_idx, up_to_round=None
            )
        except Exception as e:
            print(f"  [WARN] Feature extraction failed for {gs.filename} game {game_idx}: {e}")
            continue

        if not features_list:
            continue

        # Build a map: round_num -> features
        feat_by_round = {f.round_num: f for f in features_list}

        for round_num, round_start_ts in round_starts:
            if round_num not in feat_by_round:
                continue

            feat = feat_by_round[round_num]

            # Get model prediction
            try:
                model_prob_a = model.predict(feat)
            except Exception as e:
                continue

            # Get PM price at round START (before round resolves)
            pm_price_a = find_pm_price_at(pm_hist_sorted, round_start_ts)
            if pm_price_a is None:
                continue

            edge = model_prob_a - pm_price_a

            all_results.append(RoundResult(
                series_id=gs.series_id,
                game_idx=game_idx,
                round_num=round_num,
                model_prob_a=model_prob_a,
                pm_price_a=pm_price_a,
                edge=edge,
                team_a=team_a,
                team_b=team_b,
                team_a_won_map=team_a_won,
                map_name=map_name,
            ))

    return all_results


# ---------------------------------------------------------------------------
# P&L simulation and summary
# ---------------------------------------------------------------------------

EDGE_THRESHOLD = 0.05   # 5% minimum edge to place a bet
BET_SIZE = 10.0          # Flat $10 per bet
MIN_VOLUME = 5000        # Minimum PM market volume to include


def summarize(results: List[RoundResult], pairs: List[MatchedPair]):
    """Print backtest summary."""
    print("=" * 70)
    print("  CS2 MODEL vs POLYMARKET BACKTEST")
    print("=" * 70)
    print()

    # Matched pairs info
    unique_matches = set()
    for p in pairs:
        key = (p.grid_series.series_id, p.game_idx)
        unique_matches.add(key)
    print(f"Matched maps:             {len(unique_matches)}")
    print(f"Total round observations: {len(results)}")
    print()

    if not results:
        print("No results to analyze.")
        return

    # --- Edge distribution ---
    edges = [r.edge for r in results]
    abs_edges = [abs(e) for e in edges]
    print(f"Edge stats (model - PM):")
    print(f"  Mean edge:              {sum(edges)/len(edges):+.4f}")
    print(f"  Mean |edge|:            {sum(abs_edges)/len(abs_edges):.4f}")
    print(f"  Max edge (model higher):{max(edges):+.4f}")
    print(f"  Max edge (PM higher):   {min(edges):+.4f}")
    print()

    # --- Rounds where model found edge > threshold ---
    # We bet on the side the model favors.
    # If edge > 0, model likes team A more than PM -> bet on A
    # If edge < -threshold, model likes team B more -> bet on B
    bets = []
    for r in results:
        if r.edge > EDGE_THRESHOLD:
            # Bet on team A winning the map
            bet_won = r.team_a_won_map
            pm_price = r.pm_price_a
            model_p = r.model_prob_a
            bets.append((r, "A", bet_won, pm_price, model_p))
        elif r.edge < -EDGE_THRESHOLD:
            # Bet on team B winning the map
            bet_won = not r.team_a_won_map
            pm_price = 1.0 - r.pm_price_a
            model_p = 1.0 - r.model_prob_a
            bets.append((r, "B", bet_won, pm_price, model_p))

    print(f"Rounds with |edge| > {EDGE_THRESHOLD*100:.0f}%:  {len(bets)}")

    if not bets:
        print("No bets triggered.")
        return

    # Unique map-level bets (take the first round per map where edge > threshold)
    # In practice you'd bet once per opportunity, not every round
    map_bets = {}
    for r, side, won, pm_p, model_p in bets:
        key = (r.series_id, r.game_idx, side)
        if key not in map_bets:
            map_bets[key] = (r, side, won, pm_p, model_p)

    print(f"Unique map-level bets:    {len(map_bets)}")

    # --- Win rate ---
    wins = sum(1 for _, _, won, _, _ in bets if won)
    win_rate = wins / len(bets)
    print(f"Round-level win rate:     {wins}/{len(bets)} = {win_rate:.1%}")

    map_wins = sum(1 for _, _, won, _, _ in map_bets.values() if won)
    map_wr = map_wins / len(map_bets) if map_bets else 0
    print(f"Map-level win rate:       {map_wins}/{len(map_bets)} = {map_wr:.1%}")
    print()

    # --- P&L simulation (flat bets on PM at the available price) ---
    # On Polymarket you buy shares at price p. If outcome happens, you get $1/share.
    # Profit per share = (1 - p) if win, loss per share = -p if lose.
    # With flat $10 "risk" per bet: shares = $10 / p, profit = shares * (1-p) or loss = -$10
    # Actually, simpler: bet $10 worth of shares at price p:
    #   Win: profit = $10 * (1/p - 1) = $10 * (1-p)/p
    #   Lose: loss = -$10

    total_pnl = 0.0
    total_wagered = 0.0
    pnl_details = []

    for r, side, won, pm_p, model_p in map_bets.values():
        pm_p = max(pm_p, 0.01)  # Avoid div by zero
        if won:
            profit = BET_SIZE * (1.0 - pm_p) / pm_p
        else:
            profit = -BET_SIZE
        total_pnl += profit
        total_wagered += BET_SIZE
        pnl_details.append((r, side, won, pm_p, model_p, profit))

    print(f"--- P&L Simulation (flat ${BET_SIZE:.0f} bets, {EDGE_THRESHOLD*100:.0f}% threshold) ---")
    print(f"Total bets placed:        {len(pnl_details)}")
    print(f"Total wagered:            ${total_wagered:.2f}")
    print(f"Total P&L:                ${total_pnl:+.2f}")
    roi = (total_pnl / total_wagered * 100) if total_wagered else 0
    print(f"ROI:                      {roi:+.1f}%")
    print()

    # --- Per-bet detail ---
    print(f"{'Map':<12} {'Teams':<40} {'Side':>4} {'Model':>6} {'PM':>6} {'Edge':>7} {'Won':>4} {'P&L':>8}")
    print("-" * 90)
    for r, side, won, pm_p, model_p, profit in sorted(
        pnl_details, key=lambda x: (x[0].series_id, x[0].game_idx)
    ):
        matchup = f"{r.team_a} v {r.team_b}"
        if len(matchup) > 38:
            matchup = matchup[:38]
        print(
            f"{r.map_name:<12} {matchup:<40} {side:>4} "
            f"{model_p:>5.1%} {pm_p:>5.1%} {model_p - pm_p:>+6.1%} "
            f"{'Y' if won else 'N':>4} ${profit:>+7.2f}"
        )

    print()
    print(f"Average edge on bets:     {sum(abs(r.edge) for r, *_ in pnl_details)/len(pnl_details):.1%}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def export_excel(results: List[RoundResult], pairs: List[MatchedPair], path: str):
    """Export full round-by-round audit trail to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers

    wb = Workbook()

    # ── Sheet 1: Round-by-round detail ──────────────────────────────────
    ws = wb.active
    ws.title = "Round Detail"

    headers = [
        "Match", "Map", "Round", "Round Time (UTC)",
        "Score A", "Score B", "Team A", "Team B",
        "Model P(A)", "PM Price(A)", "Edge",
        "Bet Side", "Team A Won Map", "Bet Won",
    ]
    header_font = Font(bold=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font

    # Build score tracking per map
    map_scores = {}  # (series_id, game_idx) -> {round: (score_a, score_b)}
    for pair in pairs:
        gs = pair.grid_series
        game = gs.data["games"][pair.game_idx]
        segments = game.get("segments", [])
        score_a, score_b = 0, 0
        game_teams = game.get("teams", [])
        team_a_name = game_teams[0]["name"] if game_teams else "?"
        for seg in segments:
            seq = seg.get("sequenceNumber", 0)
            seg_teams = seg.get("teams", [])
            if len(seg_teams) >= 2:
                # Figure out which seg team is team A
                if seg_teams[0]["name"] == team_a_name:
                    if seg_teams[0].get("won"):
                        score_a += 1
                    else:
                        score_b += 1
                else:
                    if seg_teams[1].get("won"):
                        score_a += 1
                    else:
                        score_b += 1
            key = (gs.series_id, pair.game_idx)
            map_scores.setdefault(key, {})[seq] = (score_a, score_b)

    row = 2
    for r in sorted(results, key=lambda x: (x.series_id, x.game_idx, x.round_num)):
        matchup = f"{r.team_a} vs {r.team_b}"

        # Get round end time
        # Find the pair for this result to get timestamps
        round_ts = None
        for pair in pairs:
            if pair.grid_series.series_id == r.series_id and pair.game_idx == r.game_idx:
                game = pair.grid_series.data["games"][r.game_idx]
                for seg in game.get("segments", []):
                    if seg.get("sequenceNumber") == r.round_num:
                        started = seg.get("startedAt", "")
                        dur = parse_iso_duration(seg.get("duration", ""))
                        if started:
                            ts = parse_iso_ts(started) + dur
                            round_ts = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%H:%M:%S")
                        break
                break

        score_key = (r.series_id, r.game_idx)
        scores = map_scores.get(score_key, {})
        sa, sb = scores.get(r.round_num, ("?", "?"))

        edge = r.model_prob_a - r.pm_price_a
        if edge > EDGE_THRESHOLD:
            bet_side = "A"
            bet_won = r.team_a_won_map
        elif edge < -EDGE_THRESHOLD:
            bet_side = "B"
            bet_won = not r.team_a_won_map
        else:
            bet_side = ""
            bet_won = None

        row_data = [
            matchup, r.map_name, r.round_num, round_ts or "",
            sa, sb, r.team_a, r.team_b,
            r.model_prob_a, r.pm_price_a, edge,
            bet_side, "Y" if r.team_a_won_map else "N",
            ("Y" if bet_won else "N") if bet_won is not None else "",
        ]

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=c, value=val)
            # Format percentages
            if c in (9, 10, 11):
                cell.number_format = '0.0%'
            # Color edge column
            if c == 11 and isinstance(val, (int, float)):
                if abs(val) > EDGE_THRESHOLD:
                    cell.fill = green_fill if val > 0 else red_fill
            # Color bet won column
            if c == 14:
                if val == "Y":
                    cell.fill = green_fill
                elif val == "N":
                    cell.fill = red_fill

        row += 1

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

    # ── Sheet 2: Bet Summary ────────────────────────────────────────────
    ws2 = wb.create_sheet("Bet Summary")

    bet_headers = [
        "Match", "Map", "Trigger Round", "Time (UTC)", "Score",
        "Bet On", "Model Prob", "PM Price", "Edge",
        "Won", "P&L ($10 flat)",
    ]
    for c, h in enumerate(bet_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font

    # Collect unique bets (first trigger per map/side)
    map_bets = {}
    for r in sorted(results, key=lambda x: (x.series_id, x.game_idx, x.round_num)):
        edge = r.model_prob_a - r.pm_price_a
        if edge > EDGE_THRESHOLD:
            side = "A"
            pm_p = r.pm_price_a
            model_p = r.model_prob_a
            won = r.team_a_won_map
            bet_team = r.team_a
        elif edge < -EDGE_THRESHOLD:
            side = "B"
            pm_p = 1.0 - r.pm_price_a
            model_p = 1.0 - r.model_prob_a
            won = not r.team_a_won_map
            bet_team = r.team_b
        else:
            continue

        key = (r.series_id, r.game_idx, side)
        if key not in map_bets:
            # Get timestamp for this round
            round_ts = ""
            score_str = ""
            for pair in pairs:
                if pair.grid_series.series_id == r.series_id and pair.game_idx == r.game_idx:
                    game = pair.grid_series.data["games"][r.game_idx]
                    for seg in game.get("segments", []):
                        if seg.get("sequenceNumber") == r.round_num:
                            started = seg.get("startedAt", "")
                            dur = parse_iso_duration(seg.get("duration", ""))
                            if started:
                                ts = parse_iso_ts(started) + dur
                                round_ts = datetime.fromtimestamp(ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
                            break
                    break

            score_key = (r.series_id, r.game_idx)
            scores = map_scores.get(score_key, {})
            sa, sb = scores.get(r.round_num, ("?", "?"))
            score_str = f"{sa}-{sb}"

            pm_p_safe = max(pm_p, 0.01)
            pnl = BET_SIZE * (1.0 - pm_p_safe) / pm_p_safe if won else -BET_SIZE

            map_bets[key] = {
                "match": f"{r.team_a} vs {r.team_b}",
                "map": r.map_name,
                "round": r.round_num,
                "time": round_ts,
                "score": score_str,
                "bet_on": bet_team,
                "model_p": model_p,
                "pm_p": pm_p,
                "edge": model_p - pm_p,
                "won": won,
                "pnl": pnl,
            }

    row2 = 2
    total_pnl = 0
    total_bets = 0
    wins = 0
    for key in sorted(map_bets.keys()):
        b = map_bets[key]
        total_pnl += b["pnl"]
        total_bets += 1
        if b["won"]:
            wins += 1

        row_data = [
            b["match"], b["map"], b["round"], b["time"], b["score"],
            b["bet_on"], b["model_p"], b["pm_p"], b["edge"],
            "Y" if b["won"] else "N", b["pnl"],
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row2, column=c, value=val)
            if c in (7, 8, 9):
                cell.number_format = '0.0%'
            if c == 10:
                cell.fill = green_fill if val == "Y" else red_fill
            if c == 11:
                cell.number_format = '$#,##0.00'
                cell.fill = green_fill if val > 0 else red_fill

        row2 += 1

    # Summary row
    row2 += 1
    ws2.cell(row=row2, column=1, value="TOTALS").font = Font(bold=True)
    ws2.cell(row=row2, column=10, value=f"{wins}/{total_bets}").font = Font(bold=True)
    ws2.cell(row=row2, column=11, value=total_pnl).font = Font(bold=True)
    ws2.cell(row=row2, column=11).number_format = '$#,##0.00'
    row2 += 1
    ws2.cell(row=row2, column=1, value="ROI").font = Font(bold=True)
    roi = (total_pnl / (total_bets * BET_SIZE) * 100) if total_bets else 0
    ws2.cell(row=row2, column=11, value=f"{roi:+.1f}%").font = Font(bold=True)

    for col in ws2.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col_letter].width = min(max_len + 2, 35)

    wb.save(path)
    print(f"\nExported to {path}")


def main():
    print("Loading PM markets...")
    pm_markets = load_pm_markets()
    print(f"  Found {len(pm_markets)} Map Winner markets")

    print("Loading GRID series...")
    grid_series = load_grid_series()
    print(f"  Found {len(grid_series)} series")

    print("Matching GRID to PM...")
    pairs = match_grid_to_pm(grid_series, pm_markets)
    print(f"  Matched {len(pairs)} map-level pairs")

    if not pairs:
        print("\nNo matches found. Check that data/series/ and data/pm_history/ have overlapping matches.")
        return

    # Show matched pairs
    print("\n  Matched pairs:")
    for p in pairs:
        game = p.grid_series.data["games"][p.game_idx]
        gt = game.get("teams", [])
        map_name = (game.get("map") or {}).get("name", "?")
        print(f"    {p.pm_market.filename} <-> {p.grid_series.filename} "
              f"game {p.game_idx+1} ({map_name}) "
              f"[{gt[0]['name']} v {gt[1]['name']}]")

    print("\nLoading model...")
    model = MapWinModel()
    try:
        model.load()
    except FileNotFoundError as e:
        print(f"  ERROR: {e}")
        print("  Cannot run backtest without trained model.")
        return
    print(f"  Model loaded ({len(model.feature_names)} features)")

    print("\nRunning backtest...\n")
    results = run_backtest(pairs, model)

    summarize(results, pairs)

    # Export Excel
    xlsx_path = os.path.join(config.DATA_DIR, "backtest_audit.xlsx")
    export_excel(results, pairs, xlsx_path)


if __name__ == "__main__":
    main()

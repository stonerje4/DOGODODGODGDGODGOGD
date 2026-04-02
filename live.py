"""
Live paper trading system.

1. Auto-grabs PM starting odds as prior (always, before match starts)
2. Polls GRID once per round change
3. Checks real orderbook, calculates real edge
4. Paper trades: logs every BUY/SELL with timestamp + P/L
5. Detailed Excel log updated every 5 minutes + git push
6. Full round-by-round audit trail

Usage:
    python live.py --series 2912641 --pm-slug cs2-ecs-minlat-2026-03-31
    python live.py --series 2912641 --pm-slug cs2-ecs-minlat-2026-03-31 --log-dir /path/to/logs
"""

import sys, io, os, time, json, argparse, subprocess, traceback
from datetime import datetime, timezone
from dataclasses import dataclass, field, asdict
from typing import Dict, List, Optional, Tuple

if sys.platform == "win32":
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8',
                                       line_buffering=True)
    except Exception:
        pass

sys.path.insert(0, os.path.dirname(__file__))
from grid_client import GridClient
from polymarket_client import PolymarketClient
from feature_extractor import extract_features_from_grid_state
from model import MapWinModel, SeriesCalculator
from bet_sizer import BetSizer
import config


# ── Data structures ──────────────────────────────────────────────────────────

@dataclass
class Trade:
    time: str
    round_num: int
    map_num: int
    map_name: str
    market: str
    action: str       # BUY, SELL, or RESOLVE
    outcome: str
    shares: float
    price: float
    model_prob: float
    pm_price: float   # Gamma mid price at time of trade
    edge: float
    pnl: float = 0.0


@dataclass
class Position:
    market_slug: str
    market_label: str
    outcome: str
    outcome_idx: int   # 0 or 1 — resolve by index, not name
    shares: float
    avg_price: float
    entry_time: str
    entry_round: int
    entry_model_prob: float


@dataclass
class RoundSnapshot:
    """Full state snapshot per round — the audit trail."""
    time: str
    series_score: str          # "1-0"
    map_num: int
    map_name: str
    round_num: int
    score: str                 # "7-5"
    side_a: str                # "CT" or "T"
    # Model
    model_map_prob: float      # P(A wins map)
    model_series_prob: float   # P(A wins series)
    # Economy
    money_a: int
    money_b: int
    buy_type_a: str
    buy_type_b: str
    # Momentum
    kills_a: int
    kills_b: int
    first_kills_a: int
    first_kills_b: int
    streak: int
    # Market
    pm_price_a: Optional[float]       # Gamma price for team A (series winner)
    pm_ask_a: Optional[float]         # Best ask for A
    pm_bid_a: Optional[float]         # Best bid for A
    pm_liquidity: float
    # Edge
    edge_winner: Optional[float]      # model - ask - fee (series winner mkt)
    edge_map: Optional[float]         # model - ask - fee (current map mkt)
    # Positions
    open_positions: str               # Summary string
    realized_pnl: float
    unrealized_pnl: float
    # Action taken
    action: str                       # "BUY MAP1 Team A", "SELL WINNER", "HOLD", ""
    team_a: str
    team_b: str


# ── Orderbook helpers ────────────────────────────────────────────────────────

def get_book(pm, mkt):
    """Real orderbook for both sides. Returns None on any error."""
    try:
        tids = pm.get_clob_token_ids(mkt)
        if not tids or len(tids) < 2:
            return None
        out = {}
        for i, key in enumerate(["a", "b"]):
            book = pm.get_orderbook(tids[i])
            if not book:
                out[key] = {"ask": None, "bid": None, "depth": []}
                continue
            asks = sorted(book.get("asks", []), key=lambda x: float(x["price"]))
            bids = sorted(book.get("bids", []), key=lambda x: -float(x["price"]))
            depth = []
            c = 0
            for a in asks[:10]:
                p, s = float(a["price"]), float(a["size"])
                c += p * s
                depth.append({"price": p, "shares": s, "cumul": c})
            out[key] = {
                "ask": float(asks[0]["price"]) if asks else None,
                "bid": float(bids[0]["price"]) if bids else None,
                "depth": depth,
            }
        return out
    except Exception as e:
        print(f"  [WARN] Orderbook fetch failed: {e}", flush=True)
        return None


def find_edge(model_prob, book_side, fee=config.POLYMARKET_TAKER_FEE):
    """Check if edge exists and how deep."""
    ask = book_side["ask"]
    if ask is None:
        return None, 0
    edge = model_prob - ask - fee
    max_bet = 0
    for lvl in book_side["depth"]:
        if model_prob - lvl["price"] - fee > 0:
            max_bet = lvl["cumul"]
        else:
            break
    return edge, max_bet


def should_sell(model_prob, book_side, fee=config.POLYMARKET_TAKER_FEE):
    """Should we exit? True if market overvalues our position."""
    bid = book_side["bid"]
    if bid is None:
        return False, None
    return model_prob < (bid - fee), bid


# ── Team name matching ───────────────────────────────────────────────────────

def normalize_team(name):
    """Normalize for comparison."""
    return name.lower().strip().replace("team ", "").replace(" esports", "")


def teams_equivalent(name_a, name_b):
    """Check if two team names refer to the same team."""
    a, b = normalize_team(name_a), normalize_team(name_b)
    return a == b or a in b or b in a


def resolve_outcome_idx(pm_outcomes, grid_team_a, grid_team_b):
    """Figure out which PM outcome index (0 or 1) = GRID team A.
    Returns (idx_for_a, idx_for_b) or (0, 1) as fallback."""
    if len(pm_outcomes) < 2:
        return 0, 1
    for i, name in enumerate(pm_outcomes):
        if teams_equivalent(name, grid_team_a):
            return i, 1 - i
    for i, name in enumerate(pm_outcomes):
        if teams_equivalent(name, grid_team_b):
            return 1 - i, i
    return 0, 1  # fallback


# ── Excel logging ────────────────────────────────────────────────────────────

def write_excel(snapshots, trades, log_path):
    """Write full audit trail to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  [WARN] openpyxl not installed, skipping Excel export", flush=True)
        return

    wb = Workbook()

    # ── Sheet 1: Round-by-round snapshots ─────────────────────────────
    ws = wb.active
    ws.title = "Round Log"
    headers = [
        "Time (UTC)", "Series", "Map#", "Map", "Round", "Score",
        "Side A", "Model Map%", "Model Series%",
        "Money A", "Money B", "Buy A", "Buy B",
        "Kills A", "Kills B", "FK A", "FK B", "Streak",
        "PM Price A", "PM Ask A", "PM Bid A", "Liquidity",
        "Edge (Winner)", "Edge (Map)",
        "Open Positions", "Realized P/L", "Unrealized P/L",
        "Action", "Team A", "Team B",
    ]
    bold = Font(bold=True)
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h).font = bold

    for r, snap in enumerate(snapshots, 2):
        vals = [
            snap.time, snap.series_score, snap.map_num, snap.map_name,
            snap.round_num, snap.score, snap.side_a,
            snap.model_map_prob, snap.model_series_prob,
            snap.money_a, snap.money_b, snap.buy_type_a, snap.buy_type_b,
            snap.kills_a, snap.kills_b, snap.first_kills_a, snap.first_kills_b,
            snap.streak,
            snap.pm_price_a, snap.pm_ask_a, snap.pm_bid_a, snap.pm_liquidity,
            snap.edge_winner, snap.edge_map,
            snap.open_positions, snap.realized_pnl, snap.unrealized_pnl,
            snap.action, snap.team_a, snap.team_b,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (8, 9, 19, 20, 21):  # percentage columns
                cell.number_format = '0.0%'
            if c in (23, 24) and isinstance(v, (int, float)):  # edge columns
                cell.number_format = '0.0%'
                if v and v > config.MIN_EDGE_THRESHOLD:
                    cell.fill = green
                elif v and v < -config.MIN_EDGE_THRESHOLD:
                    cell.fill = red
            if c == 28 and v:  # action column
                cell.fill = yellow

    # ── Sheet 2: Trade log ────────────────────────────────────────────
    ws2 = wb.create_sheet("Trades")
    trade_headers = [
        "Time", "Round", "Map#", "Map", "Market", "Action",
        "Outcome", "Shares", "Price", "Model Prob", "PM Price",
        "Edge", "P/L",
    ]
    for c, h in enumerate(trade_headers, 1):
        ws2.cell(row=1, column=c, value=h).font = bold

    for r, t in enumerate(trades, 2):
        vals = [
            t.time, t.round_num, t.map_num, t.map_name, t.market,
            t.action, t.outcome, t.shares, t.price, t.model_prob,
            t.pm_price, t.edge, t.pnl,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            if c in (9, 10, 11, 12):
                cell.number_format = '0.000'
            if c == 13:
                cell.number_format = '$#,##0.00'
                if isinstance(v, (int, float)):
                    cell.fill = green if v > 0 else (red if v < 0 else PatternFill())
            if c == 6:  # action
                if v == "BUY":
                    cell.fill = green
                elif v == "SELL":
                    cell.fill = yellow
                elif v == "RESOLVE":
                    cell.fill = red

    # Auto-width
    for sheet in [ws, ws2]:
        for col in sheet.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    wb.save(log_path)


def git_push_log(log_dir, status_md=None):
    """Commit and push logs + status to the repo."""
    repo_root = os.path.dirname(os.path.abspath(__file__))
    try:
        if status_md:
            with open(os.path.join(repo_root, "LIVE_STATUS.md"), "w") as f:
                f.write(status_md)
        subprocess.run(["git", "add", "-A"], cwd=repo_root,
                       capture_output=True, timeout=10)
        result = subprocess.run(
            ["git", "commit", "-m",
             f"live update {datetime.now(timezone.utc).strftime('%H:%M:%S UTC')}"],
            cwd=repo_root, capture_output=True, timeout=10, text=True)
        if "nothing to commit" not in (result.stdout + result.stderr):
            subprocess.run(["git", "push", "origin", "main"],
                           cwd=repo_root, capture_output=True, timeout=30)
    except Exception as e:
        print(f"  [WARN] Git push failed: {e}", flush=True)


# ── Main ─────────────────────────────────────────────────────────────────────

def _get_starting_prior(pm, pm_slug):
    """
    Get the pre-match starting odds for team A.

    Uses PM CLOB price history: median price from 60-5 min before game start.
    Falls back to current PM price if history unavailable.

    Returns (prior_float, source_string)
    """
    import requests as _req

    try:
        mkt = pm.get_market_by_slug(pm_slug)
        if not mkt:
            return 0.50, "no_market"

        token_ids = pm.get_clob_token_ids(mkt)
        game_start_str = mkt.get("gameStartTime", "")

        if token_ids and game_start_str:
            # Parse game start time
            gs = game_start_str.replace(" ", "T")
            if not gs.endswith("Z") and "+" not in gs:
                gs += "+00:00"
            gs = gs.replace("Z", "+00:00")
            try:
                dt = datetime.fromisoformat(gs)
                start_ts = int(dt.timestamp())
                now_ts = int(datetime.now(timezone.utc).timestamp())

                if now_ts < start_ts - 300:
                    # Match hasn't started yet — grab current price, it IS the pre-match price
                    prices = pm.extract_prices(mkt)
                    outcomes = list(prices.keys())
                    if outcomes:
                        return prices[outcomes[0]], "pre_match_live"

                # Match started or about to start — pull price history
                window_start = start_ts - 3600   # 60 min before start
                window_end = start_ts - 300       # 5 min before start
                resp = _req.get(
                    f"{config.POLYMARKET_CLOB_URL}/prices-history",
                    params={
                        "market": token_ids[0],
                        "startTs": window_start,
                        "endTs": window_end,
                        "fidelity": 1,
                    },
                    timeout=15,
                )
                if resp.status_code == 200:
                    hist = resp.json().get("history", [])
                    if len(hist) >= 3:
                        prices_list = [float(h["p"]) for h in hist]
                        median_p = sorted(prices_list)[len(prices_list) // 2]
                        return median_p, f"clob_history({len(hist)}pts)"

                # Narrow window had nothing — try wider (3h before)
                window_start = start_ts - 10800
                resp = _req.get(
                    f"{config.POLYMARKET_CLOB_URL}/prices-history",
                    params={
                        "market": token_ids[0],
                        "startTs": window_start,
                        "endTs": window_end,
                        "fidelity": 1,
                    },
                    timeout=15,
                )
                if resp.status_code == 200:
                    hist = resp.json().get("history", [])
                    if hist:
                        prices_list = [float(h["p"]) for h in hist]
                        median_p = sorted(prices_list)[len(prices_list) // 2]
                        return median_p, f"clob_history_wide({len(hist)}pts)"

            except Exception as e:
                print(f"  [WARN] Price history parse error: {e}", flush=True)

        # No gameStartTime or CLOB failed — use current price as fallback
        prices = pm.extract_prices(mkt)
        outcomes = list(prices.keys())
        if outcomes:
            return prices[outcomes[0]], "current_price(no_history)"

    except Exception as e:
        print(f"  [WARN] Prior fetch failed: {e}", flush=True)

    return 0.50, "fallback_50_50"


def run(series_id, pm_slug, poll_interval=10, prior_override=None, log_dir=None):
    grid = GridClient()
    pm = PolymarketClient()
    model = MapWinModel()
    model.load()
    calc = SeriesCalculator()
    sizer = BetSizer(bankroll=1000)
    fee = config.POLYMARKET_TAKER_FEE

    # State
    positions: Dict[str, Position] = {}
    trades: List[Trade] = []
    snapshots: List[RoundSnapshot] = []
    realized_pnl = 0.0
    last_hash = None
    last_excel_write = 0

    # Logging
    if not log_dir:
        log_dir = os.path.join(config.DATA_DIR, "live_logs")
    os.makedirs(log_dir, exist_ok=True)
    log_name = f"live_{pm_slug}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    log_path = os.path.join(log_dir, log_name)

    # PM market cache: slug → market object (refreshed per-map, not per-round)
    pm_cache: Dict[str, dict] = {}
    pm_outcomes_cache: Dict[str, List[str]] = {}  # slug → [outcome_a, outcome_b]
    cache_map_key = None  # Track which map we cached for

    # ── Step 0: Get starting odds ────────────────────────────────────
    # Always try to grab PM starting odds. This is our prior.

    team_a_name = "Team A"
    team_b_name = "Team B"
    pm_idx_a = 0  # Which PM outcome index = GRID team A

    print("=" * 70, flush=True)
    print("LIVE PAPER TRADING", flush=True)
    print(f"  GRID Series: {series_id}", flush=True)
    print(f"  PM Slug:     {pm_slug}", flush=True)
    print(f"  Log:         {log_path}", flush=True)
    print("=" * 70, flush=True)

    if prior_override:
        match_prior = prior_override
        print(f"\nUsing manual prior: {match_prior:.0%}", flush=True)
    else:
        match_prior, prior_source = _get_starting_prior(pm, pm_slug)
        print(f"  Prior ({prior_source}): {match_prior:.0%}", flush=True)
        if match_prior > 0.92 or match_prior < 0.08:
            print("  WARNING: Looks mid-match - consider --prior 0.XX to override.",
                  flush=True)

    # Populate outcome names + PM cache
    try:
        mkt = pm.get_market_by_slug(pm_slug)
        if mkt:
            prices = pm.extract_prices(mkt)
            outcomes = list(prices.keys())
            pm_cache[pm_slug] = mkt
            pm_outcomes_cache[pm_slug] = outcomes
            team_a_name = outcomes[0] if outcomes else "Team A"
            team_b_name = outcomes[1] if len(outcomes) > 1 else "Team B"
            print(f"  Teams (PM): {team_a_name} vs {team_b_name}", flush=True)
    except Exception:
        pass

    if match_prior > 0.92 or match_prior < 0.08:
        print(f"\n  ⚠ Prior {match_prior:.0%} looks mid-match!", flush=True)
        print(f"  Pass --prior 0.XX to override", flush=True)

    # For future maps in Bo3: use PM map market prices if available,
    # else scale from series prior
    map_prior = 0.5 + (match_prior - 0.5) * 0.67
    print(f"  Map prior (fallback): {map_prior:.0%}", flush=True)
    print(f"\nWatching series {series_id}...\n", flush=True)

    # ── Main loop ────────────────────────────────────────────────────
    while True:
        now = datetime.now(timezone.utc).strftime("%H:%M:%S")
        now_ts = time.time()

        # ── Pull GRID (1 API call) ───────────────────────────────
        try:
            state = grid.get_series_state(series_id)
        except Exception as e:
            print(f"  [{now}] GRID error: {e}", flush=True)
            time.sleep(poll_interval)
            continue

        if not state:
            time.sleep(poll_interval)
            continue

        teams = state["teams"]
        ta, tb = teams[0]["name"], teams[1]["name"]
        sa, sb = teams[0].get("score", 0), teams[1].get("score", 0)
        games = state.get("games", [])
        is_bo3 = "3" in state.get("format", "")

        # Resolve PM outcome mapping on first GRID data
        if team_a_name == "Team A" and pm_slug in pm_outcomes_cache:
            pm_idx_a, _ = resolve_outcome_idx(
                pm_outcomes_cache[pm_slug], ta, tb)
            team_a_name = ta
            team_b_name = tb

        # Find active map
        ai = 0
        for i, gm in enumerate(games):
            if gm.get("started") and not gm.get("finished"):
                ai = i
                break
            elif gm.get("finished"):
                ai = i

        g = games[ai] if games else None
        gt = g.get("teams", []) if g else []
        mn = (g.get("map") or {}).get("name", "?") if g else "?"
        ms_a = gt[0].get("score", 0) if gt else 0
        ms_b = gt[1].get("score", 0) if len(gt) > 1 else 0
        n_segments = len(g.get("segments", [])) if g else 0

        # State change detection — includes round count
        state_hash = f"{sa}{sb}{ai}{ms_a}{ms_b}{n_segments}"
        if state_hash == last_hash:
            # No change — but still write Excel periodically
            if now_ts - last_excel_write > 300 and snapshots:
                write_excel(snapshots, trades, log_path)
                git_push_log(log_dir)
                last_excel_write = now_ts
            time.sleep(poll_interval)
            continue
        last_hash = state_hash

        # ── Refresh PM cache on map change ───────────────────────
        new_cache_key = f"{ai}_{mn}"
        if new_cache_key != cache_map_key:
            cache_map_key = new_cache_key
            pm_cache.clear()
            pm_outcomes_cache.clear()
            # Pre-fetch all market slugs for this match
            suffixes = ["", "-game1", "-game2", "-game3",
                        "-total-games-2pt5"]
            for suffix in suffixes:
                slug = pm_slug + suffix
                try:
                    mkt = pm.get_market_by_slug(slug)
                    if mkt:
                        pm_cache[slug] = mkt
                        outcomes = list(pm.extract_prices(mkt).keys())
                        pm_outcomes_cache[slug] = outcomes
                except Exception:
                    pass  # Market may not exist (e.g. game3 in a 2-0)
            print(f"  [cache] Refreshed PM markets: {list(pm_cache.keys())}",
                  flush=True)

        # ── Finished? ────────────────────────────────────────────
        if state.get("finished"):
            winner_idx = 0 if teams[0].get("won") else 1
            winner = ta if winner_idx == 0 else tb
            print(f"\n[{now}] FINISHED: {winner} wins {sa}-{sb}", flush=True)

            for slug_key, pos in list(positions.items()):
                # Resolve by outcome index — immune to name mismatches
                pm_outs = pm_outcomes_cache.get(slug_key, [])
                if pm_outs and pos.outcome_idx < len(pm_outs):
                    # For series winner: did the team at pos.outcome_idx win?
                    if pos.market_label == "WINNER":
                        won = (pos.outcome_idx == pm_idx_a and winner_idx == 0) or \
                              (pos.outcome_idx != pm_idx_a and winner_idx == 1)
                    else:
                        # Map markets: check the specific game result
                        won = _resolve_map_position(pos, games, teams)
                else:
                    # Fallback to name match
                    won = teams_equivalent(pos.outcome, winner)

                payout = 1.0 if won else 0.0
                pnl = (payout - pos.avg_price) * pos.shares
                realized_pnl += pnl
                trades.append(Trade(now, 0, ai + 1, mn, pos.market_label,
                                    "RESOLVE", pos.outcome, pos.shares,
                                    payout, 0, 0, 0, pnl))
                result = "WIN ✓" if won else "LOSS ✗"
                print(f"  {pos.market_label}: {pos.outcome} {result} "
                      f"| bought @ ${pos.avg_price:.3f} "
                      f"| P/L: ${pnl:+.2f}", flush=True)

            _print_summary(trades, realized_pnl)
            try:
                write_excel(snapshots, trades, log_path)
            except Exception as _xe:
                print(f"  [WARN] Excel write failed: {_xe}", flush=True)
            git_push_log(log_dir, _build_status_md(
                pm_slug, ta, tb, sa, sb, 0, "FINISHED", 0, 0,
                0, 0, {}, positions, trades, realized_pnl, 0, snapshots,
                finished=True))
            return

        # ── Extract features (once per game_idx, cached) ─────────
        feat_cache = {}
        for gi in range(len(games)):
            if games[gi].get("started"):
                feat_cache[gi] = extract_features_from_grid_state(
                    state, game_idx=gi)

        feats = feat_cache.get(ai, [])
        if not feats:
            time.sleep(poll_interval)
            continue

        f = feats[-1]
        p_map = model.predict(f)

        # Use PM map prices for future maps if available
        p_map2 = map_prior
        p_map3 = 0.50
        for suffix, target in [("-game2", "p_map2"), ("-game3", "p_map3")]:
            slug = pm_slug + suffix
            if slug in pm_cache:
                prices = pm.extract_prices(pm_cache[slug])
                outs = list(prices.keys())
                if outs:
                    idx_a, _ = resolve_outcome_idx(outs, ta, tb)
                    val = prices[outs[idx_a]]
                    if target == "p_map2":
                        p_map2 = val
                    else:
                        p_map3 = val

        sp = (calc.bo3_probabilities(p_map, p_map2, p_map3, sa, sb)
              if is_bo3 else calc.bo1_probabilities(p_map))

        # ── Print state ──────────────────────────────────────────
        maps_str = ""
        for i, gm in enumerate(games):
            gmt = gm.get("teams", [])
            gmn = (gm.get("map") or {}).get("name", "?")
            if gmt and gm.get("started"):
                s1, s2 = gmt[0].get("score", 0), gmt[1].get("score", 0)
                live = "<" if i == ai and not gm.get("finished") else ""
                maps_str += f"  M{i+1}({gmn}):{s1}-{s2}{live}"

        side = f.team_a_side.upper()
        print(f"[{now}] {sa}-{sb}{maps_str} R{f.round_num}({side}) "
              f"| {ta} map:{p_map:.0%} series:{sp['series_win_a']:.0%} "
              f"| ${f.money_a//1000}k/${f.money_b//1000}k "
              f"({f.buy_type_a}/{f.buy_type_b}) "
              f"K:{f.total_kills_a}-{f.total_kills_b} "
              f"FK:{f.first_kills_a}-{f.first_kills_b}", flush=True)

        # ── Check markets ────────────────────────────────────────
        # Use cached feature extractions for map probs
        def get_map_prob(idx):
            ff = feat_cache.get(idx, [])
            return model.predict(ff[-1]) if ff else None

        market_defs = [
            ("WINNER", "", sp["series_win_a"]),
            ("MAP1", "-game1", get_map_prob(0)),
            ("MAP2", "-game2", get_map_prob(1)),
            ("MAP3", "-game3", get_map_prob(2)),
            ("O/U", "-total-games-2pt5", sp["goes_to_3"]),
        ]

        # Track snapshot data for this round
        snap_pm_price = None
        snap_pm_ask = None
        snap_pm_bid = None
        snap_pm_liq = 0
        snap_edge_winner = None
        snap_edge_map = None
        snap_action = ""
        snap_unrealized = 0.0

        for label, suffix, prob_a in market_defs:
            if prob_a is None:
                continue

            slug_key = pm_slug + suffix
            mkt = pm_cache.get(slug_key)
            if not mkt:
                continue

            # Refresh prices from cache (Gamma prices are in the market obj)
            # We only fetch orderbook if there's potential edge or we hold a position
            try:
                prices = pm.extract_prices(mkt)
            except Exception:
                continue
            outcomes = list(prices.keys())
            if len(outcomes) < 2:
                continue

            # Determine which outcome is team A
            oidx_a, oidx_b = resolve_outcome_idx(outcomes, ta, tb)
            gamma_price_a = prices[outcomes[oidx_a]]

            # Track winner market price for snapshot
            if label == "WINNER":
                snap_pm_price = gamma_price_a

            # Quick edge check against Gamma mid-price before fetching orderbook
            rough_edge = abs(prob_a - gamma_price_a) - fee
            holding = slug_key in positions

            if rough_edge < 0.02 and not holding:
                # Not worth fetching orderbook — no edge, no position
                if label == "WINNER":
                    snap_edge_winner = prob_a - gamma_price_a - fee
                elif suffix.startswith("-game"):
                    snap_edge_map = prob_a - gamma_price_a - fee
                continue

            # Fetch orderbook (2 API calls)
            book = get_book(pm, mkt)
            if not book:
                continue

            liq_data = pm.extract_liquidity(mkt)
            liq = float(liq_data.get("liquidity", 0))

            if label == "WINNER":
                snap_pm_ask = book["a"]["ask"] if oidx_a == 0 else book["b"]["ask"]
                snap_pm_bid = book["a"]["bid"] if oidx_a == 0 else book["b"]["bid"]
                snap_pm_liq = liq

            # ── SELL CHECK ────────────────────────────────────
            if holding:
                pos = positions[slug_key]
                held_key = "a" if pos.outcome_idx == 0 else "b"
                held_prob = prob_a if pos.outcome_idx == oidx_a else (1 - prob_a)
                sell, bid = should_sell(held_prob, book[held_key])

                # Unrealized P/L
                if book[held_key]["bid"]:
                    unrealized = (book[held_key]["bid"] - pos.avg_price) * pos.shares
                    snap_unrealized += unrealized

                # Late-game force-exit: if map is nearly decided and we're
                # underwater, sell now rather than risk resolution at $0.
                # "Nearly decided" = one team needs 1 more round to win.
                max_score = max(ms_a, ms_b)
                force_exit = False
                if max_score >= config.ROUNDS_TO_WIN - 1 and unrealized < 0:
                    force_exit = True
                    sell = True
                    bid = book[held_key]["bid"]

                if sell and bid:
                    pnl = (bid - fee - pos.avg_price) * pos.shares
                    realized_pnl += pnl
                    gamma_mid = gamma_price_a if pos.outcome_idx == oidx_a else (1 - gamma_price_a)
                    exit_reason = "FORCE-EXIT" if force_exit else "SELL"
                    trades.append(Trade(now, f.round_num, ai + 1, mn, label,
                                        exit_reason, pos.outcome, pos.shares,
                                        bid, held_prob, gamma_mid,
                                        held_prob - bid, pnl))
                    snap_action += f"{exit_reason} {label} {pos.outcome} P/L:${pnl:+.2f}  "
                    print(f"  >>> {exit_reason} {label}: {pos.outcome} "
                          f"@ ${bid:.3f} | bought ${pos.avg_price:.3f} "
                          f"| P/L: ${pnl:+.2f} "
                          f"| model:{held_prob:.0%}"
                          f"{' (map point!)' if force_exit else ''}", flush=True)
                    del positions[slug_key]
                else:
                    if book[held_key]["bid"]:
                        mtm = (book[held_key]["bid"] - pos.avg_price) * pos.shares
                        print(f"  {label}: HOLD {pos.outcome} "
                              f"@ ${pos.avg_price:.3f} "
                              f"| mtm: ${mtm:+.2f} "
                              f"| model:{held_prob:.0%}", flush=True)
                continue

            # ── BUY CHECK ─────────────────────────────────────
            best_edge = None
            best_side = None
            for sk, oi in [("a", 0), ("b", 1)]:
                actual_oi = oidx_a if oi == 0 else oidx_b
                prob = prob_a if oi == 0 else (1 - prob_a)
                edge, max_bet = find_edge(prob, book[sk])
                if edge is not None and (best_edge is None or edge > best_edge):
                    best_edge = edge
                    best_side = {
                        "key": sk, "oi": actual_oi, "outcome": outcomes[actual_oi],
                        "prob": prob, "ask": book[sk]["ask"],
                        "edge": edge, "max_bet": max_bet,
                    }

            # Track edge for snapshot
            if label == "WINNER":
                snap_edge_winner = best_edge if best_edge else (prob_a - gamma_price_a - fee)
            elif suffix.startswith("-game"):
                snap_edge_map = best_edge if best_edge else (prob_a - gamma_price_a - fee)

            if best_side and best_side["edge"] >= config.MIN_EDGE_THRESHOLD:
                s = best_side
                # Kelly with REAL available capital
                open_exposure = sum(
                    p.shares * p.avg_price for p in positions.values())
                sizer_available = max(0, sizer.bankroll - open_exposure)
                kelly = sizer.kelly_size(s["prob"], s["ask"], liq)
                # Scale kelly by available fraction
                avail_frac = sizer_available / sizer.bankroll if sizer.bankroll > 0 else 0
                kelly_adj = kelly * avail_frac
                rec = min(kelly_adj, s["max_bet"])

                if rec >= 5:  # Min $5 bet
                    shares = rec / s["ask"]
                    positions[slug_key] = Position(
                        slug_key, label, s["outcome"], s["oi"], shares,
                        s["ask"], now, f.round_num, s["prob"])
                    gamma_mid = gamma_price_a if s["oi"] == oidx_a else (1 - gamma_price_a)
                    trades.append(Trade(now, f.round_num, ai + 1, mn, label,
                                        "BUY", s["outcome"], shares,
                                        s["ask"], s["prob"], gamma_mid,
                                        s["edge"], 0))
                    snap_action += f"BUY {label} {s['outcome']} ${rec:.0f}@{s['ask']:.3f}  "
                    print(f"  >>> BUY {label}: {s['outcome']} "
                          f"${rec:.0f} @ ${s['ask']:.3f} "
                          f"| edge:{s['edge']:+.1%} "
                          f"| depth:${s['max_bet']:.0f} "
                          f"| model:{s['prob']:.0%} "
                          f"| avail:${sizer_available:.0f}", flush=True)

        # ── P/L line ─────────────────────────────────────────
        open_count = len(positions)
        total_trades = len(trades)
        open_exposure = sum(p.shares * p.avg_price for p in positions.values())
        if open_count > 0 or total_trades > 0:
            pos_str = " | ".join(f"{p.outcome}@${p.avg_price:.2f}"
                                  for p in positions.values())
            print(f"  [P/L] realized: ${realized_pnl:+.2f} "
                  f"| unrealized: ${snap_unrealized:+.2f} "
                  f"| exposure: ${open_exposure:.0f} "
                  f"| open({open_count}): {pos_str or 'none'}", flush=True)

        # ── Record snapshot ──────────────────────────────────
        pos_summary = "; ".join(
            f"{p.market_label}:{p.outcome}@{p.avg_price:.3f}x{p.shares:.0f}"
            for p in positions.values()) or "none"

        snapshots.append(RoundSnapshot(
            time=now,
            series_score=f"{sa}-{sb}",
            map_num=ai + 1,
            map_name=mn,
            round_num=f.round_num,
            score=f"{ms_a}-{ms_b}",
            side_a=f.team_a_side.upper(),
            model_map_prob=round(p_map, 4),
            model_series_prob=round(sp["series_win_a"], 4),
            money_a=f.money_a,
            money_b=f.money_b,
            buy_type_a=f.buy_type_a,
            buy_type_b=f.buy_type_b,
            kills_a=f.total_kills_a,
            kills_b=f.total_kills_b,
            first_kills_a=f.first_kills_a,
            first_kills_b=f.first_kills_b,
            streak=f.current_streak_a,
            pm_price_a=snap_pm_price,
            pm_ask_a=snap_pm_ask,
            pm_bid_a=snap_pm_bid,
            pm_liquidity=snap_pm_liq,
            edge_winner=snap_edge_winner,
            edge_map=snap_edge_map,
            open_positions=pos_summary,
            realized_pnl=round(realized_pnl, 2),
            unrealized_pnl=round(snap_unrealized, 2),
            action=snap_action.strip(),
            team_a=ta,
            team_b=tb,
        ))

        # ── Write Excel every 5 minutes ──────────────────────
        if now_ts - last_excel_write > 300:
            try:
                write_excel(snapshots, trades, log_path)
            except Exception as _xe:
                print(f"  [WARN] Excel write failed: {_xe}", flush=True)
            git_push_log(log_dir, _build_status_md(
                pm_slug, ta, tb, sa, sb, ai, mn, ms_a, ms_b,
                f.round_num, p_map, sp, positions, trades,
                realized_pnl, snap_unrealized, snapshots))
            last_excel_write = now_ts
            print(f"  [log] Excel + LIVE_STATUS.md pushed ({len(snapshots)} rows)",
                  flush=True)

        time.sleep(poll_interval)


def _build_status_md(slug, ta, tb, sa, sb, map_num, map_name,
                     ms_a, ms_b, round_num, p_map, sp,
                     positions, trades, realized_pnl, unrealized_pnl,
                     snapshots, finished=False):
    """Build a LIVE_STATUS.md for GitHub display."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    buys = [t for t in trades if t.action == "BUY"]
    exits = [t for t in trades if t.action in ("SELL", "RESOLVE", "FORCE-EXIT")]
    wins = [t for t in exits if t.pnl > 0]
    total_risked = sum(t.shares * t.price for t in buys)
    status = "🏁 FINISHED" if finished else "🟢 LIVE"

    lines = [
        f"# CS2 Live Paper Trader — {status}",
        f"",
        f"**Last update:** {now}  ",
        f"**Match:** `{slug}`  ",
        f"**Teams:** {ta} vs {tb}  ",
        f"",
        f"## 📊 Current State",
        f"| Field | Value |",
        f"|-------|-------|",
        f"| Series Score | {ta} **{sa}** — **{sb}** {tb} |",
        f"| Map {map_num} ({map_name}) | {ms_a} — {ms_b} |",
        f"| Round | {round_num} |",
    ]
    if not finished and p_map:
        lines += [
            f"| Model P({ta} wins map) | **{p_map:.1%}** |",
            f"| Model P({ta} wins series) | **{sp.get('series_win_a', 0):.1%}** |",
            f"| P(goes to map 3) | {sp.get('goes_to_3', 0):.1%} |",
        ]

    lines += [
        f"",
        f"## 💰 P&L",
        f"| Field | Value |",
        f"|-------|-------|",
        f"| Realized P/L | **${realized_pnl:+.2f}** |",
        f"| Unrealized P/L | ${unrealized_pnl:+.2f} |",
        f"| Total risked | ${total_risked:.2f} |",
        f"| Buys | {len(buys)} |",
        f"| Exits | {len(exits)} |",
        f"| Win rate | {len(wins)}/{len(exits)} ({len(wins)/len(exits)*100:.0f}%)" if exits else "| Win rate | n/a |",
    ]

    if positions:
        lines += [f"", f"## 📋 Open Positions"]
        lines += ["| Market | Outcome | Shares | Avg Price | Entry Round |",
                  "|--------|---------|--------|-----------|-------------|"] 
        for pos in positions.values():
            lines.append(f"| {pos.market_label} | {pos.outcome} | "
                         f"{pos.shares:.0f} | ${pos.avg_price:.3f} | R{pos.entry_round} |")

    if trades:
        lines += [f"", f"## 📝 Trade Log"]
        lines += ["| Time | Market | Action | Outcome | Price | Model% | Edge | P/L |",
                  "|------|--------|--------|---------|-------|--------|------|-----|"] 
        for t in trades[-20:]:  # Last 20 trades
            pnl_str = f"${t.pnl:+.2f}" if t.action != "BUY" else "-"
            lines.append(f"| {t.time} | {t.market} | {t.action} | {t.outcome} | "
                         f"${t.price:.3f} | {t.model_prob:.0%} | {t.edge:+.1%} | {pnl_str} |")

    if snapshots:
        lines += [f"", f"## 📈 Round History (last 10)"]
        lines += ["| Time | Round | Score | Side | Map% | Series% | Money A | Money B | Buy A/B | Kills | Edge |",
                  "|------|-------|-------|------|------|---------|---------|---------|---------|-------|------|"] 
        for snap in snapshots[-10:]:
            edge_str = f"{snap.edge_winner:+.1%}" if snap.edge_winner else "-"
            lines.append(
                f"| {snap.time} | R{snap.round_num} | {snap.score} "
                f"| {snap.side_a} | {snap.model_map_prob:.0%} "
                f"| {snap.model_series_prob:.0%} "
                f"| ${snap.money_a//1000}k | ${snap.money_b//1000}k "
                f"| {snap.buy_type_a}/{snap.buy_type_b} "
                f"| {snap.kills_a}-{snap.kills_b} | {edge_str} |")

    lines += [f"", f"---", f"*Auto-updated every 5 minutes by live.py*"]
    return "\n".join(lines)


def _resolve_map_position(pos, games, teams):
    """Resolve a map-level position outcome."""
    # Parse map number from label: "MAP1" → game_idx 0
    label = pos.market_label
    if label.startswith("MAP"):
        try:
            map_idx = int(label[3:]) - 1
        except ValueError:
            return False
        if map_idx < len(games):
            game = games[map_idx]
            gt = game.get("teams", [])
            if len(gt) >= 2:
                # Check which game team matches our outcome
                for i, gteam in enumerate(gt):
                    if teams_equivalent(gteam["name"], pos.outcome):
                        return gteam.get("won", False)
    return False


def _print_summary(trades, realized_pnl):
    """End of match summary."""
    print(f"\n{'='*60}", flush=True)
    print(f"TRADE LOG:", flush=True)
    for t in trades:
        pnl_str = f"  P/L: ${t.pnl:+.2f}" if t.action in ("SELL", "RESOLVE") else ""
        print(f"  [{t.time}] {t.action:>7} {t.market:<8} {t.outcome} "
              f"x{t.shares:.0f} @ ${t.price:.3f} "
              f"(model:{t.model_prob:.0%} pm:{t.pm_price:.0%} "
              f"edge:{t.edge:+.1%}){pnl_str}", flush=True)

    buys = [t for t in trades if t.action == "BUY"]
    sells = [t for t in trades if t.action in ("SELL", "RESOLVE")]
    wins = [t for t in sells if t.pnl > 0]

    print(f"\nSUMMARY:", flush=True)
    print(f"  Total trades: {len(trades)}", flush=True)
    print(f"  Buys: {len(buys)}  |  Exits: {len(sells)}", flush=True)
    if sells:
        print(f"  Win rate: {len(wins)}/{len(sells)} "
              f"({len(wins)/len(sells)*100:.0f}%)", flush=True)
    print(f"  Realized P/L: ${realized_pnl:+.2f}", flush=True)
    total_risked = sum(t.shares * t.price for t in buys)
    if total_risked > 0:
        print(f"  Total risked: ${total_risked:.2f}", flush=True)
        print(f"  ROI: {realized_pnl/total_risked*100:+.1f}%", flush=True)
    print(f"{'='*60}", flush=True)


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="CS2 Live Paper Trading")
    p.add_argument("--series", required=True, help="GRID series ID")
    p.add_argument("--pm-slug", required=True, help="Polymarket slug")
    p.add_argument("--poll", type=int, default=10,
                   help="Poll interval seconds (default 10)")
    p.add_argument("--prior", type=float, default=None,
                   help="Manual P(Team A wins) e.g. 0.78")
    p.add_argument("--log-dir", type=str, default=None,
                   help="Directory for Excel logs (default: data/live_logs/)")
    a = p.parse_args()
    run(a.series, a.pm_slug, a.poll, a.prior, a.log_dir)

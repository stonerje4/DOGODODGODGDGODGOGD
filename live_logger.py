"""
Live match tracker with detailed Excel logging.

Logs every round snapshot: model features, predictions, market prices, edges.
Produces an Excel workbook per match with sheets:
  - Snapshots: round-by-round model vs market comparison
  - Trades: every BUY/SELL/RESOLVE with P/L
  - Summary: match overview and final stats

Usage:
    python live_logger.py --series 2917885 --pm-slug cs2-3dmax-waz-2026-04-01
    python live_logger.py --series 2917885 --pm-slug cs2-3dmax-waz-2026-04-01 --prior 0.80
    python live_logger.py --series 2917885   # GRID-only mode (no PM)
"""

import sys, io, os, time, json, argparse
from datetime import datetime, timezone
from dataclasses import dataclass, field, asdict
from typing import Dict, List, Optional, Any

if sys.platform == "win32":
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8',
                                       line_buffering=True)
    except Exception:
        pass

sys.path.insert(0, os.path.dirname(__file__))
from grid_client import GridClient
from polymarket_client import PolymarketClient
from feature_extractor import extract_features_from_grid_state, features_to_dict
from model import MapWinModel, SeriesCalculator
from bet_sizer import BetSizer
import config

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Excel Writer ─────────────────────────────────────────────────────────────

class MatchLogger:
    """Writes detailed round-by-round data to Excel."""

    def __init__(self, series_id: str, team_a: str, team_b: str):
        self.series_id = series_id
        self.team_a = team_a
        self.team_b = team_b

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_a = team_a.replace(" ", "_")[:12]
        safe_b = team_b.replace(" ", "_")[:12]
        self.filename = os.path.join(
            os.path.dirname(__file__),
            f"log_{safe_a}_vs_{safe_b}_{ts}.xlsx"
        )

        self.wb = openpyxl.Workbook()

        # ── Snapshots sheet ──────────────────────────────────
        self.ws_snap = self.wb.active
        self.ws_snap.title = "Snapshots"
        self.snap_headers = [
            "Timestamp", "Map#", "Map", "Round",
            "Score_A", "Score_B", "Score_Diff",
            "Side_A", "Is_2nd_Half",
            # Economy
            "Money_A", "Money_B", "Money_Diff",
            "Buy_A", "Buy_B",
            "Loss_Tier_A", "Loss_Tier_B",
            # Momentum
            "Last3_A", "Last5_A", "Streak_A",
            # Skill
            "Kills_A", "Kills_B", "Kill_Diff",
            "FK_A", "FK_B",
            # Pistol
            "Pistol1", "Pistol2",
            # Model outputs
            "Model_Map_Win_A", "Model_Series_Win_A",
            "Model_Goes_To_3", "Model_A_Sweeps", "Model_B_Sweeps",
            # Market prices (PM)
            "PM_Winner_Price_A", "PM_Map_Price_A",
            "PM_Winner_Ask_A", "PM_Winner_Bid_A",
            "PM_Map_Ask_A", "PM_Map_Bid_A",
            # Edge
            "Edge_Winner", "Edge_Map",
            # Bet action
            "Action", "Action_Detail",
        ]
        self.ws_snap.append(self.snap_headers)
        self._style_header(self.ws_snap, len(self.snap_headers))

        # ── Trades sheet ─────────────────────────────────────
        self.ws_trades = self.wb.create_sheet("Trades")
        self.trade_headers = [
            "Timestamp", "Market", "Action", "Outcome",
            "Shares", "Price", "Model_Prob", "Market_Price",
            "Edge", "Kelly_Size", "P/L",
        ]
        self.ws_trades.append(self.trade_headers)
        self._style_header(self.ws_trades, len(self.trade_headers))

        # ── Summary sheet ────────────────────────────────────
        self.ws_summary = self.wb.create_sheet("Summary")
        self.summary_data = {}

        self.snap_row = 2
        self.trade_row = 2

    def _style_header(self, ws, ncols):
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496",
                               fill_type="solid")
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

    def log_snapshot(self, snap: dict):
        """Write one row to Snapshots sheet."""
        row = [snap.get(h, "") for h in self.snap_headers]
        self.ws_snap.append(row)

        # Color-code edge columns
        r = self.ws_snap.max_row
        for col_name in ["Edge_Winner", "Edge_Map"]:
            if col_name in self.snap_headers:
                ci = self.snap_headers.index(col_name) + 1
                cell = self.ws_snap.cell(row=r, column=ci)
                try:
                    val = float(cell.value) if cell.value else 0
                    if val > 0.08:
                        cell.fill = PatternFill(start_color="C6EFCE",
                                                fill_type="solid")
                    elif val > 0:
                        cell.fill = PatternFill(start_color="FFEB9C",
                                                fill_type="solid")
                    elif val < -0.05:
                        cell.fill = PatternFill(start_color="FFC7CE",
                                                fill_type="solid")
                except (ValueError, TypeError):
                    pass

        # Color-code action column
        if "Action" in self.snap_headers:
            ci = self.snap_headers.index("Action") + 1
            cell = self.ws_snap.cell(row=r, column=ci)
            if cell.value == "BUY":
                cell.font = Font(bold=True, color="006100")
            elif cell.value == "SELL":
                cell.font = Font(bold=True, color="9C0006")

        self.snap_row += 1

    def log_trade(self, trade: dict):
        """Write one row to Trades sheet."""
        row = [trade.get(h, "") for h in self.trade_headers]
        self.ws_trades.append(row)

        r = self.ws_trades.max_row
        action_ci = self.trade_headers.index("Action") + 1
        cell = self.ws_trades.cell(row=r, column=action_ci)
        if cell.value == "BUY":
            cell.font = Font(bold=True, color="006100")
        elif cell.value in ("SELL", "RESOLVE"):
            cell.font = Font(bold=True, color="9C0006")

        # P/L coloring
        pnl_ci = self.trade_headers.index("P/L") + 1
        pnl_cell = self.ws_trades.cell(row=r, column=pnl_ci)
        try:
            pnl_val = float(pnl_cell.value) if pnl_cell.value else 0
            if pnl_val > 0:
                pnl_cell.font = Font(bold=True, color="006100")
            elif pnl_val < 0:
                pnl_cell.font = Font(bold=True, color="9C0006")
        except (ValueError, TypeError):
            pass

        self.trade_row += 1

    def write_summary(self, data: dict):
        """Write summary sheet at end of match."""
        ws = self.ws_summary
        title_font = Font(bold=True, size=14)
        label_font = Font(bold=True)

        ws.cell(row=1, column=1, value=f"{self.team_a} vs {self.team_b}").font = title_font
        ws.cell(row=2, column=1, value=f"Series ID: {self.series_id}")
        ws.cell(row=3, column=1, value=f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}")

        row = 5
        for key, val in data.items():
            ws.cell(row=row, column=1, value=key).font = label_font
            ws.cell(row=row, column=2, value=val)
            row += 1

    def save(self):
        """Auto-fit columns and save."""
        for ws in [self.ws_snap, self.ws_trades, self.ws_summary]:
            for col_cells in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 3, 25)

        self.wb.save(self.filename)
        return self.filename


# ── Orderbook helpers (same as live.py) ──────────────────────────────────────

def get_book(pm, mkt):
    tids = pm.get_clob_token_ids(mkt)
    if not tids or len(tids) < 2:
        return None
    out = {}
    for i, key in enumerate(["a", "b"]):
        book = pm.get_orderbook(tids[i])
        if not book:
            out[key] = {"ask": None, "bid": None}
            continue
        asks = sorted(book.get("asks", []), key=lambda x: float(x["price"]))
        bids = sorted(book.get("bids", []), key=lambda x: -float(x["price"]))
        out[key] = {
            "ask": float(asks[0]["price"]) if asks else None,
            "bid": float(bids[0]["price"]) if bids else None,
        }
    return out


# ── Main loop ────────────────────────────────────────────────────────────────

def run(series_id, pm_slug=None, poll_interval=10, prior_override=None):
    grid = GridClient()
    pm = PolymarketClient() if pm_slug else None
    model = MapWinModel()
    model.load()
    calc = SeriesCalculator()
    sizer = BetSizer(bankroll=1000)
    fee = config.POLYMARKET_TAKER_FEE

    positions: Dict[str, dict] = {}
    trades: List[dict] = []
    realized_pnl = 0.0
    last_hash = None
    logger = None  # Created once we know team names
    snap_count = 0

    # Pre-match prior
    match_prior = prior_override or 0.50
    if pm_slug and not prior_override:
        print("Grabbing pre-match prior from PM...", flush=True)
        mkt = pm.get_market_by_slug(pm_slug)
        if mkt:
            prices = pm.extract_prices(mkt)
            outcomes = list(prices.keys())
            if outcomes:
                match_prior = prices[outcomes[0]]
        print(f"Prior: {match_prior:.0%}", flush=True)

    map_prior = 0.5 + (match_prior - 0.5) * 0.67

    print(f"Watching series {series_id}...", flush=True)
    if pm_slug:
        print(f"PM slug: {pm_slug}", flush=True)
    else:
        print("Running GRID-only mode (no Polymarket)", flush=True)
    print(flush=True)

    while True:
        now_ts = datetime.now(timezone.utc).strftime("%H:%M:%S")

        state = grid.get_series_state(series_id)
        if not state:
            print(f"  [{now_ts}] Waiting for GRID data...", flush=True)
            time.sleep(poll_interval)
            continue

        t = state["teams"]
        ta, tb = t[0]["name"], t[1]["name"]
        sa, sb = t[0].get("score", 0), t[1].get("score", 0)
        games = state.get("games", [])
        is_bo3 = "3" in state.get("format", "")

        # Create logger on first data
        if logger is None:
            logger = MatchLogger(series_id, ta, tb)
            print(f"Logging to: {logger.filename}", flush=True)
            print(flush=True)

        # Find active map
        ai = 0
        for i, g in enumerate(games):
            if g.get("started") and not g.get("finished"):
                ai = i
                break
            elif g.get("finished"):
                ai = i

        g = games[ai] if games else None
        gt = g.get("teams", []) if g else []
        mn = (g.get("map") or {}).get("name", "?") if g else "?"
        ms_a = gt[0].get("score", 0) if gt else 0
        ms_b = gt[1].get("score", 0) if len(gt) > 1 else 0

        state_hash = f"{sa}{sb}{ai}{ms_a}{ms_b}"
        if state_hash == last_hash:
            time.sleep(poll_interval)
            continue
        last_hash = state_hash

        # ── FINISHED ─────────────────────────────────────────
        if state.get("finished"):
            winner = ta if t[0].get("won") else tb
            print(f"\n[{now_ts}] FINISHED: {winner} wins {sa}-{sb}", flush=True)

            # Resolve positions
            for slug_key, pos in list(positions.items()):
                won = pos["outcome"] == winner
                payout = 1.0 if won else 0.0
                pnl = (payout - pos["avg_price"]) * pos["shares"]
                realized_pnl += pnl
                result = "WIN" if won else "LOSS"

                trade_rec = {
                    "Timestamp": now_ts, "Market": pos.get("label", slug_key),
                    "Action": "RESOLVE", "Outcome": pos["outcome"],
                    "Shares": round(pos["shares"], 2),
                    "Price": payout, "Model_Prob": "",
                    "Market_Price": "", "Edge": "",
                    "Kelly_Size": "", "P/L": round(pnl, 2),
                }
                trades.append(trade_rec)
                logger.log_trade(trade_rec)
                print(f"  {pos.get('label','?')}: {pos['outcome']} {result} "
                      f"| bought @ ${pos['avg_price']:.3f} "
                      f"| P/L: ${pnl:+.2f}", flush=True)

            # Write summary
            buys = [t for t in trades if t["Action"] == "BUY"]
            exits = [t for t in trades if t["Action"] in ("SELL", "RESOLVE")]
            wins = [t for t in exits if (t.get("P/L") or 0) > 0]
            logger.write_summary({
                "Winner": winner,
                "Final Score": f"{sa}-{sb}",
                "Format": state.get("format", "?"),
                "Total Snapshots": snap_count,
                "Total Trades": len(trades),
                "Buys": len(buys),
                "Exits": len(exits),
                "Win Rate": f"{len(wins)}/{len(exits)} ({len(wins)/max(len(exits),1)*100:.0f}%)" if exits else "N/A",
                "Realized P/L": f"${realized_pnl:+.2f}",
                "Pre-match Prior": f"{match_prior:.0%}",
            })

            saved = logger.save()
            print(f"\nLog saved: {saved}", flush=True)
            return

        # ── Extract features ─────────────────────────────────
        feats = extract_features_from_grid_state(state, game_idx=ai)
        if not feats:
            time.sleep(poll_interval)
            continue

        f = feats[-1]
        p_map = model.predict(f)
        sp = (calc.bo3_probabilities(p_map, map_prior, 0.50, sa, sb)
              if is_bo3 else calc.bo1_probabilities(p_map))

        # ── Market data ──────────────────────────────────────
        pm_winner_price_a = None
        pm_map_price_a = None
        pm_winner_book = None
        pm_map_book = None
        action = ""
        action_detail = ""

        if pm:
            # Winner market
            mkt_w = pm.get_market_by_slug(pm_slug)
            if mkt_w:
                prices_w = pm.extract_prices(mkt_w)
                outcomes_w = list(prices_w.keys())
                if outcomes_w:
                    pm_winner_price_a = prices_w[outcomes_w[0]]
                pm_winner_book = get_book(pm, mkt_w)

            # Current map market
            map_suffix = f"-game{ai+1}"
            mkt_m = pm.get_market_by_slug(pm_slug + map_suffix)
            if mkt_m:
                prices_m = pm.extract_prices(mkt_m)
                outcomes_m = list(prices_m.keys())
                if outcomes_m:
                    pm_map_price_a = prices_m[outcomes_m[0]]
                pm_map_book = get_book(pm, mkt_m)

            # ── Check for trades (winner market) ─────────────
            if mkt_w and pm_winner_book:
                slug_key = pm_slug
                if slug_key in positions:
                    pos = positions[slug_key]
                    held_side = "a" if pos["outcome"] == outcomes_w[0] else "b"
                    held_prob = sp["series_win_a"] if held_side == "a" else (1 - sp["series_win_a"])
                    bid = pm_winner_book[held_side]["bid"]
                    if bid and held_prob < (bid - fee):
                        pnl = (bid - fee - pos["avg_price"]) * pos["shares"]
                        realized_pnl += pnl
                        action = "SELL"
                        action_detail = f"SELL {pos['outcome']} @ ${bid:.3f} P/L ${pnl:+.2f}"
                        trade_rec = {
                            "Timestamp": now_ts, "Market": "WINNER",
                            "Action": "SELL", "Outcome": pos["outcome"],
                            "Shares": round(pos["shares"], 2),
                            "Price": round(bid, 4),
                            "Model_Prob": round(held_prob, 4),
                            "Market_Price": round(bid, 4),
                            "Edge": round(held_prob - bid, 4),
                            "Kelly_Size": "", "P/L": round(pnl, 2),
                        }
                        trades.append(trade_rec)
                        logger.log_trade(trade_rec)
                        del positions[slug_key]
                elif slug_key not in positions:
                    # Check buy
                    best = None
                    for sk, oi in [("a", 0), ("b", 1)]:
                        prob = sp["series_win_a"] if oi == 0 else (1 - sp["series_win_a"])
                        ask = pm_winner_book[sk]["ask"]
                        if ask is None:
                            continue
                        edge = prob - ask - fee
                        if best is None or edge > best["edge"]:
                            best = {"outcome": outcomes_w[oi], "prob": prob,
                                    "ask": ask, "edge": edge, "side": sk}

                    if best and best["edge"] > 0:
                        liq = float(pm.extract_liquidity(mkt_w).get("liquidity", 0))
                        kelly = sizer.kelly_size(best["prob"], best["ask"], liq)
                        if kelly >= 5:
                            shares = kelly / best["ask"]
                            positions[slug_key] = {
                                "outcome": best["outcome"], "shares": shares,
                                "avg_price": best["ask"], "entry_time": now_ts,
                                "label": "WINNER",
                            }
                            action = "BUY"
                            action_detail = (f"BUY {best['outcome']} ${kelly:.0f} "
                                             f"@ ${best['ask']:.3f} edge:{best['edge']:+.1%}")
                            trade_rec = {
                                "Timestamp": now_ts, "Market": "WINNER",
                                "Action": "BUY", "Outcome": best["outcome"],
                                "Shares": round(shares, 2),
                                "Price": round(best["ask"], 4),
                                "Model_Prob": round(best["prob"], 4),
                                "Market_Price": round(best["ask"], 4),
                                "Edge": round(best["edge"], 4),
                                "Kelly_Size": round(kelly, 2),
                                "P/L": 0,
                            }
                            trades.append(trade_rec)
                            logger.log_trade(trade_rec)

        # ── Compute edges ────────────────────────────────────
        edge_winner = None
        edge_map = None
        if pm_winner_price_a is not None:
            edge_winner = round(sp["series_win_a"] - pm_winner_price_a, 4)
        if pm_map_price_a is not None:
            edge_map = round(p_map - pm_map_price_a, 4)

        # ── Build snapshot row ───────────────────────────────
        snap = {
            "Timestamp": now_ts,
            "Map#": ai + 1,
            "Map": mn,
            "Round": f.round_num,
            "Score_A": f.score_a,
            "Score_B": f.score_b,
            "Score_Diff": f.score_diff,
            "Side_A": f.team_a_side.upper(),
            "Is_2nd_Half": "Y" if f.is_second_half else "N",
            "Money_A": f.money_a,
            "Money_B": f.money_b,
            "Money_Diff": f.money_diff,
            "Buy_A": f.buy_type_a,
            "Buy_B": f.buy_type_b,
            "Loss_Tier_A": f.loss_tier_a,
            "Loss_Tier_B": f.loss_tier_b,
            "Last3_A": f.last_3_wins_a,
            "Last5_A": f.last_5_wins_a,
            "Streak_A": f.current_streak_a,
            "Kills_A": f.total_kills_a,
            "Kills_B": f.total_kills_b,
            "Kill_Diff": f.total_kills_a - f.total_kills_b,
            "FK_A": f.first_kills_a,
            "FK_B": f.first_kills_b,
            "Pistol1": f.pistol_1_winner,
            "Pistol2": f.pistol_2_winner,
            "Model_Map_Win_A": round(p_map, 4),
            "Model_Series_Win_A": round(sp["series_win_a"], 4),
            "Model_Goes_To_3": round(sp["goes_to_3"], 4),
            "Model_A_Sweeps": round(sp["a_sweeps"], 4),
            "Model_B_Sweeps": round(sp["b_sweeps"], 4),
            "PM_Winner_Price_A": round(pm_winner_price_a, 4) if pm_winner_price_a is not None else "",
            "PM_Map_Price_A": round(pm_map_price_a, 4) if pm_map_price_a is not None else "",
            "PM_Winner_Ask_A": round(pm_winner_book["a"]["ask"], 4) if pm_winner_book and pm_winner_book["a"]["ask"] else "",
            "PM_Winner_Bid_A": round(pm_winner_book["a"]["bid"], 4) if pm_winner_book and pm_winner_book["a"]["bid"] else "",
            "PM_Map_Ask_A": round(pm_map_book["a"]["ask"], 4) if pm_map_book and pm_map_book["a"]["ask"] else "",
            "PM_Map_Bid_A": round(pm_map_book["a"]["bid"], 4) if pm_map_book and pm_map_book["a"]["bid"] else "",
            "Edge_Winner": edge_winner if edge_winner is not None else "",
            "Edge_Map": edge_map if edge_map is not None else "",
            "Action": action,
            "Action_Detail": action_detail,
        }

        logger.log_snapshot(snap)
        snap_count += 1

        # Auto-save every snapshot so we never lose data
        logger.save()

        # ── Console output ───────────────────────────────────
        maps_str = ""
        for i, gm in enumerate(games):
            gmt = gm.get("teams", [])
            gmn = (gm.get("map") or {}).get("name", "?")
            if gmt and gm.get("started"):
                s1, s2 = gmt[0].get("score", 0), gmt[1].get("score", 0)
                live = "<" if i == ai and not gm.get("finished") else ""
                maps_str += f"  M{i+1}({gmn}):{s1}-{s2}{live}"

        side = f.team_a_side.upper()
        pm_str = ""
        if pm_winner_price_a is not None:
            pm_str = f" | PM:{pm_winner_price_a:.0%}"
        if edge_winner is not None:
            pm_str += f" edge:{edge_winner:+.0%}"

        print(f"[{now_ts}] {sa}-{sb}{maps_str} R{f.round_num}({side}) "
              f"| {ta} map:{p_map:.0%} series:{sp['series_win_a']:.0%}"
              f"{pm_str}"
              f" | econ:{f.buy_type_a}/{f.buy_type_b} "
              f"${f.money_a:,}/${f.money_b:,} "
              f"K:{f.total_kills_a}-{f.total_kills_b}",
              flush=True)

        if action:
            print(f"  >>> {action_detail}", flush=True)

        # P/L line
        if positions or trades:
            pos_str = " | ".join(f"{p['outcome']}@${p['avg_price']:.2f}"
                                  for p in positions.values())
            print(f"  [P/L] realized: ${realized_pnl:+.2f} "
                  f"| open: {pos_str or 'none'} "
                  f"| trades: {len(trades)} | snaps: {snap_count}",
                  flush=True)

        time.sleep(poll_interval)


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Live match tracker with Excel logging")
    p.add_argument("--series", required=True, help="GRID series ID")
    p.add_argument("--pm-slug", default=None,
                   help="Polymarket market slug (optional, runs GRID-only if omitted)")
    p.add_argument("--poll", type=int, default=10,
                   help="Poll interval seconds (default 10)")
    p.add_argument("--prior", type=float, default=None,
                   help="Pre-match P(Team A wins) override")
    a = p.parse_args()
    run(a.series, a.pm_slug, a.poll, a.prior)

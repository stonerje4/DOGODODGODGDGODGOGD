"""
Live match logger — polls GRID + Polymarket, runs model, logs every round to Excel.

One row per round. Every detail: scores, economy, model prob, PM price, edge, signals.

Usage:
    python live_logger_v2.py --series 2917927 --pm-slug cs2-cw-5s-2026-04-02 --prior 0.65 --team-a "Chinggis Warriors"
"""

import time
import argparse
import os
import json
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import config
from grid_client import GridClient
from polymarket_client import PolymarketClient
from feature_extractor import extract_features_from_grid_state, features_to_dict
from model import MapWinModel, SeriesCalculator
from bet_sizer import BetSizer


# ── Styling ──────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
SIGNAL_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
MAP_BREAK_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)


COLUMNS = [
    # Timing
    ("Timestamp", 14),
    ("Map #", 6),
    ("Map Name", 10),
    ("Round", 7),
    # Scores
    ("Score A", 8),
    ("Score B", 8),
    ("Round Winner", 14),
    ("Series Score", 11),
    # Sides & Economy
    ("A Side", 7),
    ("A Money", 10),
    ("B Money", 10),
    ("A Buy", 7),
    ("B Buy", 7),
    ("A Loss Tier", 9),
    ("B Loss Tier", 9),
    # Momentum
    ("Last 3 (A)", 9),
    ("Last 5 (A)", 9),
    ("Streak", 7),
    ("Kill Diff", 8),
    # Model Output
    ("Model P(A Map)", 13),
    ("Model P(A Series)", 14),
    ("P(Goes to 3)", 11),
    # Polymarket
    ("PM Price A", 10),
    ("PM Price B", 10),
    ("PM Volume", 10),
    ("PM Liquidity", 11),
    # Edge
    ("Raw Edge", 9),
    ("Eff Edge", 9),
    ("Signal?", 8),
    ("Bet Side", 10),
    ("Kelly Size $", 10),
    # Notes
    ("Notes", 30),
]


class LiveLoggerV2:
    def __init__(
        self,
        grid_series_id: str,
        pm_slug: str,
        prior_a: float,
        team_a_label: str,
        poll_interval: int = 12,
        output_dir: str = None,
    ):
        self.grid = GridClient()
        self.pm = PolymarketClient()
        self.model = MapWinModel()
        self.model.load()
        self.series_calc = SeriesCalculator()
        self.bet_sizer = BetSizer()

        self.grid_series_id = grid_series_id
        self.pm_slug = pm_slug
        self.prior_a = prior_a
        self.team_a_label = team_a_label
        self.poll_interval = poll_interval

        # State tracking
        self.last_round_key = (0, 0)  # (game_idx, round_num)
        self.rounds_logged = 0

        # Excel setup
        output_dir = output_dir or os.path.dirname(__file__)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_slug = pm_slug.replace("/", "_")
        self.xlsx_path = os.path.join(output_dir, f"log_{safe_slug}_{ts}.xlsx")
        self._init_excel()

    def _init_excel(self):
        """Create workbook with formatted header."""
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "Live Rounds"

        # Info rows at top
        self.ws.merge_cells("A1:F1")
        self.ws["A1"] = f"Match: {self.team_a_label} | GRID: {self.grid_series_id} | PM: {self.pm_slug}"
        self.ws["A1"].font = Font(bold=True, size=12)
        self.ws.merge_cells("A2:F2")
        self.ws["A2"] = f"Prior: {self.team_a_label} = {self.prior_a:.0%} | Started: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        self.ws["A2"].font = Font(size=10, italic=True)

        # Headers at row 4
        for col_idx, (name, width) in enumerate(COLUMNS, 1):
            cell = self.ws.cell(row=4, column=col_idx, value=name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            self.ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        self.ws.freeze_panes = "A5"
        self.data_start_row = 5
        self._save()

    def _save(self):
        try:
            self.wb.save(self.xlsx_path)
        except PermissionError:
            # File might be open in Excel, try alternate name
            alt = self.xlsx_path.replace(".xlsx", "_v2.xlsx")
            self.wb.save(alt)

    def _write_row(self, row_data: dict, is_signal: bool = False, is_map_break: bool = False):
        """Write one row to Excel."""
        row_num = self.data_start_row + self.rounds_logged

        for col_idx, (col_name, _) in enumerate(COLUMNS, 1):
            val = row_data.get(col_name, "")
            cell = self.ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

            # Format percentages
            if col_name in ("Model P(A Map)", "Model P(A Series)", "P(Goes to 3)",
                            "PM Price A", "PM Price B", "Raw Edge", "Eff Edge"):
                if isinstance(val, (int, float)):
                    cell.number_format = "0.0%"

            if col_name in ("A Money", "B Money", "PM Volume", "PM Liquidity", "Kelly Size $"):
                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0"

        if is_signal:
            for col_idx in range(1, len(COLUMNS) + 1):
                self.ws.cell(row=row_num, column=col_idx).fill = SIGNAL_FILL
        elif is_map_break:
            for col_idx in range(1, len(COLUMNS) + 1):
                self.ws.cell(row=row_num, column=col_idx).fill = MAP_BREAK_FILL

        self.rounds_logged += 1
        self._save()

    def run(self):
        """Main polling loop."""
        print(f"{'=' * 70}")
        print(f"LIVE LOGGER V2")
        print(f"  GRID: {self.grid_series_id}")
        print(f"  PM:   {self.pm_slug}")
        print(f"  Prior: {self.team_a_label} = {self.prior_a:.0%}")
        print(f"  Excel: {self.xlsx_path}")
        print(f"  Poll:  every {self.poll_interval}s")
        print(f"{'=' * 70}")

        prev_game_idx = -1

        while True:
            try:
                now_str = datetime.now(timezone.utc).strftime("%H:%M:%S")

                # ── Pull GRID ────────────────────────────────────────────
                grid_state = self.grid.get_series_state(self.grid_series_id)
                if not grid_state:
                    print(f"  [{now_str}] Waiting for GRID data...")
                    time.sleep(self.poll_interval)
                    continue

                if grid_state.get("finished"):
                    print(f"\n  [{now_str}] SERIES FINISHED")
                    self._write_summary()
                    break

                if not grid_state.get("started"):
                    print(f"  [{now_str}] Match not started yet...")
                    time.sleep(self.poll_interval)
                    continue

                games = grid_state.get("games", [])
                if not games:
                    time.sleep(self.poll_interval)
                    continue

                # Find current active game
                current_game_idx = 0
                for i, g in enumerate(games):
                    if g.get("started") and not g.get("finished"):
                        current_game_idx = i
                        break
                    elif g.get("finished"):
                        current_game_idx = i

                game = games[current_game_idx]
                segments = game.get("segments", [])
                finished_rounds = [s for s in segments if s.get("finished")]
                current_round = len(finished_rounds)

                # Check if new round
                round_key = (current_game_idx, current_round)
                if round_key == self.last_round_key:
                    time.sleep(self.poll_interval)
                    continue

                # Map break row
                if current_game_idx != prev_game_idx and prev_game_idx >= 0:
                    prev_g = games[prev_game_idx]
                    prev_teams = prev_g.get("teams", [])
                    prev_map = (prev_g.get("map") or {}).get("name", "?")
                    winner = "?"
                    for pt in prev_teams:
                        if pt.get("won"):
                            winner = pt.get("name", "?")
                    self._write_row({
                        "Timestamp": now_str,
                        "Notes": f"=== MAP {prev_game_idx+1} FINISHED ({prev_map}) — Winner: {winner} ===",
                    }, is_map_break=True)

                prev_game_idx = current_game_idx
                self.last_round_key = round_key

                if current_round == 0:
                    time.sleep(self.poll_interval)
                    continue

                # ── Series score ─────────────────────────────────────────
                series_teams = grid_state.get("teams", [])
                team_a_name = series_teams[0].get("name", "?") if series_teams else "?"
                team_b_name = series_teams[1].get("name", "?") if len(series_teams) > 1 else "?"
                s_score_a = series_teams[0].get("score", 0) if series_teams else 0
                s_score_b = series_teams[1].get("score", 0) if len(series_teams) > 1 else 0

                # ── Map info ─────────────────────────────────────────────
                map_name = (game.get("map") or {}).get("name", "?")
                game_teams = game.get("teams", [])
                score_a = game_teams[0].get("score", 0) if game_teams else 0
                score_b = game_teams[1].get("score", 0) if len(game_teams) > 1 else 0

                # Last round winner
                last_seg = finished_rounds[-1] if finished_rounds else None
                round_winner = "?"
                if last_seg:
                    for t in last_seg.get("teams", []):
                        if t.get("won"):
                            round_winner = t.get("name", "?")

                # ── Run model ────────────────────────────────────────────
                features_list = extract_features_from_grid_state(
                    grid_state, game_idx=current_game_idx)

                if not features_list:
                    print(f"  [{now_str}] No features yet")
                    time.sleep(self.poll_interval)
                    continue

                latest = features_list[-1]
                p_map = self.model.predict(latest)

                # Series calc — use prior for future maps
                p_map2_est = self.prior_a
                p_map3_est = 0.50 + (self.prior_a - 0.50) * 0.5

                fmt = grid_state.get("format", "best-of-3")
                if "1" in fmt:
                    series_probs = self.series_calc.bo1_probabilities(p_map)
                else:
                    series_probs = self.series_calc.bo3_probabilities(
                        p_current_map=p_map,
                        p_map2=p_map2_est,
                        p_map3=p_map3_est,
                        series_score_a=s_score_a,
                        series_score_b=s_score_b,
                    )

                p_series = series_probs["series_win_a"]
                p_goes3 = series_probs["goes_to_3"]

                # ── Pull PM prices ───────────────────────────────────────
                pm_price_a = None
                pm_price_b = None
                pm_vol = 0
                pm_liq = 0

                try:
                    pm_market = self.pm.get_market_by_slug(self.pm_slug)
                    if pm_market:
                        prices = self.pm.extract_prices(pm_market)
                        liq_info = self.pm.extract_liquidity(pm_market)
                        outcomes = list(prices.keys())
                        if len(outcomes) >= 2:
                            pm_price_a = prices[outcomes[0]]
                            pm_price_b = prices[outcomes[1]]
                        pm_vol = liq_info.get("volume", 0)
                        pm_liq = liq_info.get("liquidity", 0)
                except Exception as e:
                    pass  # PM might be slow, that's fine

                # ── Edge calc ────────────────────────────────────────────
                raw_edge = 0
                eff_edge = 0
                is_signal = False
                bet_side = ""
                kelly = 0
                notes = ""

                if pm_price_a is not None:
                    # Edge on team A side
                    edge_a = p_series - pm_price_a
                    # Edge on team B side
                    edge_b = (1 - p_series) - pm_price_b

                    if abs(edge_a) >= abs(edge_b):
                        raw_edge = edge_a
                        if edge_a > 0:
                            bet_side = team_a_name
                            eff_edge = edge_a - config.POLYMARKET_TAKER_FEE
                        else:
                            bet_side = team_b_name
                            raw_edge = -edge_a
                            eff_edge = -edge_a - config.POLYMARKET_TAKER_FEE
                    else:
                        raw_edge = abs(edge_b)
                        if edge_b > 0:
                            bet_side = team_b_name
                            eff_edge = edge_b - config.POLYMARKET_TAKER_FEE
                        else:
                            bet_side = team_a_name
                            eff_edge = -edge_b - config.POLYMARKET_TAKER_FEE

                    if eff_edge > config.MIN_EDGE_THRESHOLD:
                        is_signal = True
                        ask = pm_price_a if bet_side == team_a_name else pm_price_b
                        kelly = self.bet_sizer.kelly_size(
                            prob=p_series if bet_side == team_a_name else (1 - p_series),
                            ask_price=ask,
                            liquidity=pm_liq,
                        )
                        notes = f"SIGNAL: {bet_side} @ {ask:.2f}"

                    # Add context notes
                    if latest.buy_type_a == "eco":
                        notes += " | A on ECO"
                    if latest.buy_type_b == "eco":
                        notes += " | B on ECO"
                    if current_round == 1:
                        notes += " | PISTOL"
                    if current_round == config.ROUNDS_PER_HALF + 1:
                        notes += " | 2ND HALF PISTOL"
                    if current_round == config.ROUNDS_PER_HALF:
                        notes += " | LAST ROUND 1ST HALF"
                else:
                    notes = "PM data unavailable"

                # ── Log row ──────────────────────────────────────────────
                row = {
                    "Timestamp": now_str,
                    "Map #": current_game_idx + 1,
                    "Map Name": map_name,
                    "Round": current_round,
                    "Score A": score_a,
                    "Score B": score_b,
                    "Round Winner": round_winner,
                    "Series Score": f"{s_score_a}-{s_score_b}",
                    "A Side": latest.team_a_side.upper(),
                    "A Money": latest.money_a,
                    "B Money": latest.money_b,
                    "A Buy": latest.buy_type_a,
                    "B Buy": latest.buy_type_b,
                    "A Loss Tier": latest.loss_tier_a,
                    "B Loss Tier": latest.loss_tier_b,
                    "Last 3 (A)": latest.last_3_wins_a,
                    "Last 5 (A)": latest.last_5_wins_a,
                    "Streak": latest.current_streak_a,
                    "Kill Diff": latest.total_kills_a - latest.total_kills_b,
                    "Model P(A Map)": round(p_map, 4),
                    "Model P(A Series)": round(p_series, 4),
                    "P(Goes to 3)": round(p_goes3, 4),
                    "PM Price A": pm_price_a,
                    "PM Price B": pm_price_b,
                    "PM Volume": pm_vol,
                    "PM Liquidity": pm_liq,
                    "Raw Edge": round(raw_edge, 4) if raw_edge else 0,
                    "Eff Edge": round(eff_edge, 4) if eff_edge else 0,
                    "Signal?": "YES" if is_signal else "",
                    "Bet Side": bet_side if is_signal else "",
                    "Kelly Size $": kelly if is_signal else "",
                    "Notes": notes.strip(" |"),
                }

                self._write_row(row, is_signal=is_signal)

                # Console output
                sig_tag = " *** SIGNAL ***" if is_signal else ""
                print(f"  [{now_str}] Map{current_game_idx+1} R{current_round} "
                      f"{score_a}-{score_b} ({map_name}) | "
                      f"Model:{p_map:.1%} Series:{p_series:.1%} | "
                      f"PM:{pm_price_a:.2f}/{pm_price_b:.2f} | "
                      f"Edge:{eff_edge:+.1%}{sig_tag}")

            except KeyboardInterrupt:
                print("\nStopped.")
                break
            except Exception as e:
                print(f"  ERROR: {e}")
                import traceback
                traceback.print_exc()

            time.sleep(self.poll_interval)

        self._save()
        print(f"\n  Excel saved: {self.xlsx_path}")

    def _write_summary(self):
        """Write summary sheet."""
        ws2 = self.wb.create_sheet("Summary")
        ws2["A1"] = "Match Summary"
        ws2["A1"].font = Font(bold=True, size=14)

        ws2["A3"] = "Total Rounds Logged"
        ws2["B3"] = self.rounds_logged
        ws2["A4"] = "Signals Found"
        ws2["B4"] = sum(1 for r in range(self.data_start_row, self.data_start_row + self.rounds_logged)
                        if self.ws.cell(row=r, column=29).value == "YES")
        ws2["A5"] = "Prior (Team A)"
        ws2["B5"] = self.prior_a
        ws2["A6"] = "GRID Series ID"
        ws2["B6"] = self.grid_series_id
        ws2["A7"] = "PM Slug"
        ws2["B7"] = self.pm_slug

        self._save()


def main():
    parser = argparse.ArgumentParser(description="Live match logger to Excel")
    parser.add_argument("--series", required=True, help="GRID series ID")
    parser.add_argument("--pm-slug", required=True, help="Polymarket slug (base, no date)")
    parser.add_argument("--prior", type=float, required=True, help="Pre-match P(Team A) from PM (e.g. 0.65)")
    parser.add_argument("--team-a", type=str, default="Team A", help="Team A display name")
    parser.add_argument("--interval", type=int, default=12, help="Poll interval seconds")
    args = parser.parse_args()

    logger = LiveLoggerV2(
        grid_series_id=args.series,
        pm_slug=args.pm_slug,
        prior_a=args.prior,
        team_a_label=args.team_a,
        poll_interval=args.interval,
    )
    logger.run()


if __name__ == "__main__":
    main()
